from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import OuterRef, Subquery, Exists, Count, Value, IntegerField, Q
from django.db.models.functions import Coalesce
from django.utils import timezone
from django.core.cache import cache
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.db import transaction
from datetime import datetime, timedelta, date
import csv
import io
import re
import openpyxl
from io import BytesIO
import pandas as pd
import logging
import threading
from decimal import Decimal
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import subprocess
from .models import Device, Reading, SyncStatus, MeterModel, Region, Substation, Feeder, TransformerSubstation, DeviceGroup
from django.core.paginator import Paginator
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


logger = logging.getLogger(__name__)

# ===== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ =====
def is_allowed_user(user):
    return user.is_staff or user.groups.filter(name='operator').exists()

def get_producer_display(code, type_id):
    """Преобразует код производителя и ID типа в человекочитаемое название."""
    if code == "SX":
        if type_id == "18":
            return "Sanxing_new"
        elif type_id == "22":
            return "Sanxing_old"
        else:
            return "Sanxing (unknown)"
    mapping = {
        "RS": "RiseSun",
        "EM": "Energomera",
        "SR": "SunRise",
        "UK": "Hexing KUK",
        "ST": "Star",
        "HX": "Hexing",
        "8": "Others",
    }
    return mapping.get(code, code or "Неизвестно")

def normalize_model_name(model):
    """Приводит название модели к каноническому виду для сравнения."""
    if not model:
        return ''
    # Удаляем лишние пробелы, приводим к нижнему регистру, убираем точки/дефисы/пробелы
    return re.sub(r'[^a-z0-9]', '', model.lower().strip())

def compare_models(model_1c, model_db):
    """
    Возвращает тип совпадения: 'exact', 'partial', 'none'
    """
    if not model_1c or not model_db:
        return 'none'
    norm_1c = normalize_model_name(model_1c)
    norm_db = normalize_model_name(model_db)
    if norm_1c == norm_db:
        return 'exact'
    if norm_1c in norm_db or norm_db in norm_1c:
        return 'partial'
    # Дополнительно: удаляем суффиксы типа -S31.543 и сравниваем
    # Удаляем всё после дефиса, если есть
    base_1c = norm_1c.split('-')[0] if '-' in norm_1c else norm_1c
    base_db = norm_db.split('-')[0] if '-' in norm_db else norm_db
    if base_1c == base_db and len(base_1c) > 3:
        return 'partial'
    return 'none'

def suggest_correct_serial(serial_1c, device_found=False):
    """
    Предлагает правильный вариант заводского номера для ввода в 1С.
    Если устройство не найдено, но номер длиннее 8 цифр, предлагаем последние 8 цифр.
    """
    if not serial_1c:
        return None
    # Удаляем все нецифровые символы
    digits = re.sub(r'\D', '', serial_1c)
    if len(digits) > 8:
        return digits[-8:]
    return None

def find_device(serial_number, model=None):
    """
    Поиск устройства по серийному номеру.
    Сначала точное совпадение, затем по суффиксу (последние символы).
    Возвращает объект Device или None.
    """
    if not serial_number:
        return None
    sn = str(serial_number).strip()
    qs = Device.objects.filter(serial_number=sn)
    if model:
        qs = qs.filter(model=model)
    if qs.exists():
        return qs.first()
    candidates = Device.objects.filter(serial_number__endswith=sn)
    if model:
        candidates = candidates.filter(model=model)
    if candidates.count() == 1:
        return candidates.first()
    return None

# ===== АВТОРИЗАЦИЯ =====
def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Неверный логин или пароль')
    return render(request, 'meters/login.html')

def logout_view(request):
    logout(request)
    return redirect('dashboard')

# ===== ДАШБОРД =====
def dashboard(request):
    period = request.GET.get('period', 'today')
    cache_key = f'dashboard_stats_{period}'
    data = cache.get(cache_key)
    if data is None or request.GET.get('refresh') == '1':
        today = timezone.now().date()
        if period == 'today':
            start_date = today
        elif period == '3d':
            start_date = today - timedelta(days=3)
        elif period == '7d':
            start_date = today - timedelta(days=7)
        elif period == '30d':
            start_date = today - timedelta(days=30)
        else:
            start_date = today

        start_dt = datetime.combine(start_date, datetime.min.time())
        end_dt = datetime.combine(today, datetime.max.time())

        total_devices = Device.objects.filter(status='active').count()

        active_device_ids = Reading.objects.filter(
            device__status='active',
            timestamp__gte=start_dt,
            timestamp__lte=end_dt
        ).values_list('device_id', flat=True).distinct()
        active_today = active_device_ids.count()
        active_percent = round((active_today / total_devices * 100), 1) if total_devices else 0

        auto_notes = [
            "Авто-сбор: База cEnergo",
            "Авто-сбор: Sanxing_old",
            "Авто-сбор: SunRise",
            "Hexing KUK",
            "RiseSun"
        ]
        manual_today = Reading.objects.filter(
            device__status='active',
            timestamp__gte=start_dt,
            timestamp__lte=end_dt
        ).exclude(notes__in=auto_notes).count()

        devices = Device.objects.filter(status='active').select_related('model').only(
            'id', 'model__device_type_str', 'model__device_type_id'
        )
        active_set = set(active_device_ids)
        producer_stats_dict = {}
        for device in devices:
            model = device.model
            if model:
                code = model.device_type_str or ""
                type_id = model.device_type_id or ""
                display_name = get_producer_display(code, type_id)
            else:
                display_name = "Неизвестно"
            if display_name not in producer_stats_dict:
                producer_stats_dict[display_name] = {'total': 0, 'active': 0}
            producer_stats_dict[display_name]['total'] += 1
            if device.id in active_set:
                producer_stats_dict[display_name]['active'] += 1

        producer_stats = []
        for name, stats in producer_stats_dict.items():
            percent = round((stats['active'] / stats['total'] * 100), 1) if stats['total'] else 0
            producer_stats.append({
                'producer': name,
                'total': stats['total'],
                'active': stats['active'],
                'percent': percent,
            })
        producer_stats.sort(key=lambda x: x['producer'])

        data = {
            'total_devices': total_devices,
            'active_today': active_today,
            'active_percent': active_percent,
            'manual_today': manual_today,
            'sync_statuses': list(SyncStatus.objects.all().order_by('-last_update')),
            'producer_stats': producer_stats,
            'period': period,
        }
        cache.set(cache_key, data, 600)
    return render(request, 'meters/dashboard.html', data)

# ===== УСТРОЙСТВА (Реестр) – ОБНОВЛЁННАЯ ВЕРСИЯ =====
@login_required
@user_passes_test(is_allowed_user)
def devices(request):
    # Получаем фильтры из GET
    search = request.GET.get('search', '')
    producer_filter = request.GET.get('producer', '')
    model_filter = request.GET.get('model', '')
    status_filter = request.GET.get('status', '')
    region_filter = request.GET.get('region', '')
    substation_filter = request.GET.get('substation', '')
    feeder_filter = request.GET.get('feeder', '')
    tp_filter = request.GET.get('tp', '')
    group_filter = request.GET.get('group', '')
    category_filter = request.GET.get('category', '')

    # Все группы для выпадающих списков (в модалку)
    groups_all = DeviceGroup.objects.all().order_by('name')

    # Пагинация
    page = int(request.GET.get('page', 1))
    per_page = 10

    # Базовый запрос с предзагрузкой связей
    qs = Device.objects.select_related('model', 'region', 'substation', 'feeder', 'tp').prefetch_related('groups')

    # ===== ЭКСПОРТ В EXCEL =====
    # if request.GET.get('export') == 'xlsx':
    #     import openpyxl
    #     from openpyxl.styles import Font, Alignment, PatternFill
    #     from openpyxl.utils import get_column_letter

    #     # Создаём книгу и лист
    #     wb = openpyxl.Workbook()
    #     ws = wb.active
    #     ws.title = "Реестр приборов"

    #     # Заголовки
    #     headers = [
    #         'Заводской номер', 'Тип ПУ', 'Статус', 'Район', 'Подстанция',
    #         'Фидер', 'ТП', 'Группы', 'Номинальный ток', 'Фазность', 'Дата добавления'
    #     ]
    #     # Заполняем заголовки
    #     for col, header in enumerate(headers, 1):
    #         cell = ws.cell(row=1, column=col, value=header)
    #         cell.font = Font(bold=True, color='FFFFFF')
    #         cell.fill = PatternFill(start_color='2d3e50', end_color='2d3e50', fill_type='solid')
    #         cell.alignment = Alignment(horizontal='center')

    #     # Получаем данные с учётом фильтров (qs уже отфильтрован)
    #     # Для оптимизации подгружаем связанные модели
    #     devices_export = qs.select_related('model', 'region', 'substation', 'feeder', 'tp').prefetch_related('groups')

    #     for row_idx, device in enumerate(devices_export, 2):
    #         # Группы в виде строки через запятую
    #         groups = ', '.join(g.name for g in device.groups.all())
    #         ws.cell(row=row_idx, column=1, value=device.serial_number)
    #         ws.cell(row=row_idx, column=2, value=device.model.model_name if device.model else '')
    #         ws.cell(row=row_idx, column=3, value=device.status)
    #         ws.cell(row=row_idx, column=4, value=device.region.name if device.region else '')
    #         ws.cell(row=row_idx, column=5, value=device.substation.name if device.substation else '')
    #         ws.cell(row=row_idx, column=6, value=device.feeder.name if device.feeder else '')
    #         ws.cell(row=row_idx, column=7, value=device.tp.name if device.tp else '')
    #         ws.cell(row=row_idx, column=8, value=groups)
    #         ws.cell(row=row_idx, column=9, value=device.nominal_current or '')
    #         ws.cell(row=row_idx, column=10, value=device.model.phases if device.model else '')
    #         ws.cell(row=row_idx, column=11, value=device.created_at.strftime('%Y-%m-%d %H:%M:%S') if device.created_at else '')

    #     # Автоширина колонок
    #     for col in ws.columns:
    #         max_length = 0
    #         col_letter = get_column_letter(col[0].column)
    #         for cell in col:
    #             try:
    #                 if cell.value:
    #                     max_length = max(max_length, len(str(cell.value)))
    #             except:
    #                 pass
    #         adjusted_width = max_length + 2
    #         ws.column_dimensions[col_letter].width = min(adjusted_width, 50)

    #     response = HttpResponse(
    #         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    #     )
    #     response['Content-Disposition'] = f'attachment; filename="registry_devices_{timezone.now().date()}.xlsx"'
    #     wb.save(response)
    #     return response

    # Применяем фильтры
    if search:
        qs = qs.filter(Q(serial_number__icontains=search) | Q(model__model_name__icontains=search))
    if producer_filter:
        qs = qs.filter(model__device_type_str=producer_filter)
    if model_filter:
        qs = qs.filter(model__id=model_filter)
    if status_filter:
        qs = qs.filter(status=status_filter)
    if region_filter:
        qs = qs.filter(region_id=region_filter)
    if substation_filter:
        qs = qs.filter(substation_id=substation_filter)
    if feeder_filter:
        qs = qs.filter(feeder_id=feeder_filter)
    if tp_filter:
        qs = qs.filter(tp_id=tp_filter)
    if group_filter:
        qs = qs.filter(groups__id=group_filter)
    if category_filter == '__none__':
        qs = qs.filter(category__isnull=True)
    elif category_filter:
        qs = qs.filter(category=category_filter)

    from django.db.models import Exists, OuterRef
    from django.utils import timezone
    from datetime import timedelta
    seven_days_ago = timezone.now() - timedelta(days=7)
    qs = qs.annotate(
        has_recent_reading=Exists(
            Reading.objects.filter(
                device=OuterRef('pk'),
                timestamp__gte=seven_days_ago
            )
        )
    )

    total = qs.count()
    total_pages = (total + per_page - 1) // per_page if total > 0 else 1
    if page > total_pages:
        page = total_pages
    devices_list = qs.order_by('serial_number')[(page-1)*per_page:page*per_page]
    page_range = range(1, total_pages + 1)

    # Получаем списки для выпадающих фильтров
    codes = MeterModel.objects.values_list('device_type_str', flat=True).distinct().order_by('device_type_str')
    producers_with_selected = []
    for code in codes:
        display = get_producer_display(code, '')
        if display == "Неизвестно" and code:
            display = code
        selected = (code == producer_filter)
        producers_with_selected.append((display, code, selected))

    models_list = MeterModel.objects.all().order_by('model_name')
    models_with_selected = [(m, str(m.id) == model_filter) for m in models_list]

    statuses = [
        ('active', 'Активен'),
        ('repair', 'Ремонт'),
        ('inactive', 'Отключен'),
        ('temporary_off', 'Временно отключен'),
    ]
    statuses_with_selected = [(value, label, value == status_filter) for value, label in statuses]

    # Списки для иерархии
    regions_with_selected = [(r, str(r.id) == region_filter) for r in Region.objects.all().order_by('name')]
    substations_with_selected = [(s, str(s.id) == substation_filter) for s in Substation.objects.all().order_by('name')]
    feeders_with_selected = [(f, str(f.id) == feeder_filter) for f in Feeder.objects.all().order_by('name')]
    tps_with_selected = [(t, str(t.id) == tp_filter) for t in TransformerSubstation.objects.all().order_by('name')]
    groups_with_selected = [(g, str(g.id) == group_filter) for g in DeviceGroup.objects.all().order_by('name')]

    categories = list(Device.CATEGORY_CHOICES)
    categories_with_selected = [(c[0], c[1], c[0] == category_filter) for c in categories]
    categories_with_selected.insert(0, ('', 'Все', category_filter == ''))
    categories_with_selected.append(('__none__', 'Не указано', category_filter == '__none__'))

    # Обработка POST-запроса
    if request.method == 'POST':
        # Массовое обновление привязки к иерархии (существующий код)
        if 'update_hierarchy' in request.POST:
            selected_ids = request.POST.getlist('selected_devices')
            if not selected_ids:
                messages.warning(request, 'Выберите хотя бы одно устройство.')
                return redirect(request.path)

            target_region = request.POST.get('target_region')
            target_substation = request.POST.get('target_substation')
            target_feeder = request.POST.get('target_feeder')
            target_tp = request.POST.get('target_tp')

            update_fields = {}
            if target_region and target_region != '':
                update_fields['region_id'] = target_region
            if target_substation and target_substation != '':
                update_fields['substation_id'] = target_substation
            if target_feeder and target_feeder != '':
                update_fields['feeder_id'] = target_feeder
            if target_tp and target_tp != '':
                update_fields['tp_id'] = target_tp

            if not update_fields:
                messages.warning(request, 'Выберите хотя бы одно поле для обновления.')
                return redirect(request.path)

            updated_count = Device.objects.filter(id__in=selected_ids).update(**update_fields)
            messages.success(request, f'Обновлено {updated_count} устройств.')
            return redirect(request.path)

        # НОВОЕ: массовое добавление устройств в группы через модалку
        elif 'add_to_groups_modal' in request.POST:
            selected_ids = request.POST.getlist('selected_devices')
            group_ids = request.POST.getlist('group_ids')

            if not selected_ids:
                messages.warning(request, 'Выберите хотя бы одно устройство.')
            elif not group_ids:
                messages.warning(request, 'Выберите хотя бы одну группу.')
            else:
                devices = Device.objects.filter(id__in=selected_ids)
                groups_to_add = DeviceGroup.objects.filter(id__in=group_ids)
                for device in devices:
                    device.groups.add(*groups_to_add)
                messages.success(request, f'Устройства добавлены в {len(groups_to_add)} группу(ы).')
            return redirect(request.path)
        
        # НОВОЕ: массовое изменение устройств категории
        elif 'update_category' in request.POST:
            selected_ids = request.POST.getlist('selected_devices')
            new_category = request.POST.get('new_category')
            if not selected_ids:
                messages.warning(request, 'Выберите хотя бы одно устройство.')
            elif not new_category:
                messages.warning(request, 'Выберите категорию.')
            else:
                Device.objects.filter(id__in=selected_ids).update(category=new_category)
                messages.success(request, f'Категория обновлена для {len(selected_ids)} устройств.')
            return redirect(request.path)

    context = {
        'devices': devices_list,
        'total': total,
        'page': page,
        'per_page': per_page,
        'search': search,
        'total_pages': total_pages,
        'page_range': page_range,
        'producers_with_selected': producers_with_selected,
        'models_with_selected': models_with_selected,
        'statuses_with_selected': statuses_with_selected,
        'producer_filter': producer_filter,
        'model_filter': model_filter,
        'status_filter': status_filter,
        # 'regions': regions,
        # 'substations': substations,
        # 'feeders': feeders,
        # 'tps': tps,
        'groups_all': groups_all,
        'regions_with_selected': regions_with_selected,
        'substations_with_selected': substations_with_selected,
        'feeders_with_selected': feeders_with_selected,
        'tps_with_selected': tps_with_selected,
        'groups_with_selected': groups_with_selected,                     
        'selected_region': region_filter,
        'selected_substation': substation_filter,
        'selected_feeder': feeder_filter,
        'selected_tp': tp_filter,
        'selected_group': group_filter,
        'categories_with_selected': categories_with_selected,
        'category_filter': category_filter,
    }

    current_params = request.GET.copy()
    if 'page' in current_params:
        current_params.pop('page')
    current_params = current_params.urlencode()
    context['current_params'] = current_params

    return render(request, 'meters/devices.html', context)

# ===== УПРАВЛЕНИЕ ГРУППАМИ ДЛЯ ОТДЕЛЬНОГО УСТРОЙСТВА =====
@login_required
@user_passes_test(is_allowed_user)
def device_add_to_groups(request, device_id):
    device = get_object_or_404(Device, id=device_id)
    if request.method == 'POST':
        group_ids = request.POST.getlist('group_ids')
        if group_ids:
            groups = DeviceGroup.objects.filter(id__in=group_ids)
            device.groups.add(*groups)
            messages.success(request, f'Устройство {device.serial_number} добавлено в выбранные группы.')
        else:
            messages.warning(request, 'Не выбрано ни одной группы.')
        return redirect('devices')
    groups = DeviceGroup.objects.all().order_by('name')
    return render(request, 'meters/device_add_to_groups.html', {'device': device, 'groups': groups})

@login_required
@user_passes_test(is_allowed_user)
def device_remove_from_groups(request, device_id):
    device = get_object_or_404(Device, id=device_id)
    if request.method == 'POST':
        group_ids = request.POST.getlist('group_ids')
        if group_ids:
            groups = DeviceGroup.objects.filter(id__in=group_ids)
            device.groups.remove(*groups)
            messages.success(request, f'Устройство {device.serial_number} удалено из выбранных групп.')
        else:
            messages.warning(request, 'Не выбрано ни одной группы.')
        return redirect('devices')
    groups = device.groups.all().order_by('name')
    return render(request, 'meters/device_remove_from_groups.html', {'device': device, 'groups': groups})

@login_required
@user_passes_test(is_allowed_user)
def edit_device(request, serial_number):
    device = get_object_or_404(Device, serial_number=serial_number)
    if request.method == 'POST':
        new_serial = request.POST.get('serial_number', '').strip()
        if not new_serial:
            messages.error(request, 'Номер не может быть пустым.')
        elif new_serial == device.serial_number:
            messages.warning(request, 'Номер не изменён.')
        elif Device.objects.filter(serial_number=new_serial).exclude(id=device.id).exists():
            messages.error(request, f'Номер "{new_serial}" уже используется другим устройством.')
        else:
            device.serial_number = new_serial
            device.save()
            messages.success(request, f'Заводской номер изменён на {new_serial}.')
            return redirect('devices')
    context = {'device': device}
    return render(request, 'meters/edit_device.html', context)

@login_required
@user_passes_test(is_allowed_user)
def delete_device(request, serial_number):
    device = get_object_or_404(Device, serial_number=serial_number)
    if request.method == 'POST':
        device.delete()
        messages.success(request, f'Устройство {serial_number} удалено.')
        return redirect('devices')
    return render(request, 'meters/confirm_delete.html', {'object': device, 'type': 'устройство'})

@login_required
@user_passes_test(is_allowed_user)
def add_device(request):
    if request.method == 'POST':
        if 'single_submit' in request.POST:
            serial = request.POST.get('serial_number', '').strip()
            model_id = request.POST.get('model')
            status = request.POST.get('status', 'active')
            if not serial:
                messages.error(request, 'Серийный номер обязателен.')
            elif Device.objects.filter(serial_number=serial).exists():
                messages.error(request, f'Прибор с номером {serial} уже существует.')
            else:
                try:
                    model = MeterModel.objects.get(id=model_id) if model_id else None
                    if model:
                        nominal_current = model.nominal_current
                        phase = model.phases
                        askue_id = model.device_type_id or ''
                        api_id = model.device_type_str or ''
                    else:
                        nominal_current = ''
                        phase = None
                        askue_id = ''
                        api_id = ''
                    jpes_region, _ = Region.objects.get_or_create(name='ЖПЭС')
                    Device.objects.create(
                        serial_number=serial,
                        model=model,
                        status=status,
                        region=jpes_region,
                        nominal_current=nominal_current,
                        phase=phase,
                        askue_id=askue_id,
                        api_id=api_id,
                    )
                    messages.success(request, f'Прибор {serial} добавлен.')
                    return redirect('devices')
                except Exception as e:
                    messages.error(request, f'Ошибка: {e}')
        elif 'bulk_submit' in request.POST:
            serials_text = request.POST.get('bulk_serials', '').strip()
            model_id = request.POST.get('bulk_model')
            status = request.POST.get('bulk_status', 'active')
            if not serials_text:
                messages.error(request, 'Введите хотя бы один номер.')
            else:
                serials = [s.strip() for s in serials_text.split('\n') if s.strip()]
                added = 0
                skipped = []
                try:
                    model = MeterModel.objects.get(id=model_id) if model_id else None
                    if model:
                        nominal_current = model.nominal_current
                        phase = model.phases
                        askue_id = model.device_type_id or ''
                        api_id = model.device_type_str or ''
                    else:
                        nominal_current = ''
                        phase = None
                        askue_id = ''
                        api_id = ''
                    for sn in serials:
                        if Device.objects.filter(serial_number=sn).exists():
                            skipped.append(sn)
                            continue
                        Device.objects.create(
                            serial_number=sn,
                            model=model,
                            status=status,
                            nominal_current=nominal_current,
                            phase=phase,
                            askue_id=askue_id,
                            api_id=api_id,
                        )
                        added += 1
                except Exception as e:
                    skipped.append(f"Ошибка: {e}")
                messages.success(request, f'Добавлено: {added}. Пропущено: {len(skipped)}.')
                if skipped:
                    messages.warning(request, f'Пропущенные: {", ".join(skipped[:5])}')
                return redirect('devices')
    context = {
        'models': MeterModel.objects.all().order_by('model_name'),
    }
    return render(request, 'meters/add_device.html', context)

@login_required
@user_passes_test(is_allowed_user)
def upload_devices(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        file = request.FILES['excel_file']
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            added = 0
            skipped = []
            errors = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                sn = str(row[0]).strip() if row[0] else None
                model_name = str(row[1]).strip() if row[1] else None
                status = str(row[2]).strip() if row[2] else 'active'
                nominal_current = str(row[3]).strip() if row[3] else ''
                phase = row[4]
                askue_id = str(row[5]).strip() if row[5] else ''
                api_id = str(row[6]).strip() if row[6] else ''
                if not sn:
                    continue
                if Device.objects.filter(serial_number=sn).exists():
                    skipped.append(sn)
                    continue
                try:
                    model = MeterModel.objects.filter(model_name=model_name).first() if model_name else None
                    jpes_region, _ = Region.objects.get_or_create(name='ЖПЭС')
                    Device.objects.create(
                        serial_number=sn,
                        model=model,
                        status=status,
                        region=jpes_region,
                        nominal_current=nominal_current,
                        phase=int(phase) if phase else None,
                        askue_id=askue_id,
                        api_id=api_id,
                    )
                    added += 1
                except Exception as e:
                    errors.append(f"{sn}: {e}")
            messages.success(request, f'Добавлено: {added}. Пропущено (дубликаты): {len(skipped)}. Ошибок: {len(errors)}')
            if errors:
                messages.warning(request, f'Ошибки: {"; ".join(errors[:3])}')
        except Exception as e:
            messages.error(request, f'Ошибка обработки файла: {e}')
        return redirect('devices')
    return redirect('add_device')

@login_required
@user_passes_test(is_allowed_user)
def download_template(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Devices"
    headers = ["serial_number", "model_name", "status", "nominal_current", "phase", "askue_id", "api_id"]
    ws.append(headers)
    ws.append(["084600001431", "DTZY217", "active", "5(80)", "3", "", ""])
    ws.append(["084600001432", "P34S02", "active", "10(80)", "3", "", ""])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=device_template.xlsx'
    wb.save(response)
    return response

# ===== ПОКАЗАНИЯ =====
def readings(request):
    serial = request.GET.get('serial', '')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    device = None
    readings_list = []
    is_online = False
    last_reading_date = None

    today = timezone.now().date()
    start_3d = (today - timedelta(days=3)).strftime('%Y-%m-%d')
    start_7d = (today - timedelta(days=7)).strftime('%Y-%m-%d')
    start_30d = (today - timedelta(days=30)).strftime('%Y-%m-%d')
    end_today = today.strftime('%Y-%m-%d')

    if serial:
        device = Device.objects.filter(serial_number=serial).first()
        if device:
            qs = Reading.objects.filter(device=device)
            if start_date_str:
                try:
                    start_dt = datetime.strptime(start_date_str, '%Y-%m-%d')
                    qs = qs.filter(timestamp__gte=start_dt)
                except:
                    pass
            if end_date_str:
                try:
                    end_dt = datetime.strptime(end_date_str, '%Y-%m-%d') + timedelta(days=1)
                    qs = qs.filter(timestamp__lt=end_dt)
                except:
                    pass

            if not start_date_str and not end_date_str:
                readings_list = qs.order_by('-timestamp')[:10]
            else:
                readings_list = qs.order_by('-timestamp')

            last_reading = Reading.objects.filter(device=device).order_by('-timestamp').first()
            if last_reading:
                last_reading_date = last_reading.timestamp
                days_since_last = (timezone.now() - last_reading.timestamp).days
                is_online = days_since_last < 7

    context = {
        'device': device,
        'readings': readings_list,
        'serial': serial,
        'is_online': is_online,
        'last_reading_date': last_reading_date,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'today': today,
        'start_3d': start_3d,
        'start_7d': start_7d,
        'start_30d': start_30d,
        'end_today': end_today,
    }
    return render(request, 'meters/readings.html', context)

# ===== ГРУППОВОЙ ВВОД И ИМПОРТ EXCEL (С ФОНОМ) =====
def run_import_background(session_key):
    from django.contrib.sessions.models import Session
    try:
        session = Session.objects.get(session_key=session_key)
        data = session.get_decoded()
        import_data = data.get('import_data', [])
        params = data.get('import_params', {})
        if not import_data or not params:
            logger.warning("Нет данных для фонового импорта")
            return

        col_serial = params.get('col_serial')
        col_timestamp = params.get('col_timestamp')
        col_value = params.get('col_value')
        col_notes = params.get('col_notes')
        date_format = params.get('date_format', 'auto')
        col_direction = params.get('col_direction')
        default_direction = params.get('default_direction', 'aplus')

        if col_direction and row.get(col_direction):
            direction = str(row[col_direction]).strip().lower()
            if direction in ['a-', 'aminus']:
                direction = 'aminus'
            else:
                direction = 'aplus'
        else:
            direction = default_direction

        success = 0
        errors = []
        skipped = []
        skipped_empty = []
        updated = 0

        with transaction.atomic():
            for row in import_data:
                sn = str(row.get(col_serial, '')).strip()
                if not sn:
                    continue
                device = find_device(sn)
                if not device:
                    skipped.append(sn)
                    continue

                raw_value = row.get(col_value)
                if pd.isna(raw_value) or raw_value == '':
                    skipped_empty.append(sn)
                    continue
                try:
                    val = float(raw_value)
                except:
                    errors.append(f"Ошибка преобразования значения для {sn}: {raw_value}")
                    continue

                raw_date = row.get(col_timestamp)
                if pd.isna(raw_date):
                    continue

                try:
                    if isinstance(raw_date, datetime):
                        dt = raw_date
                    elif isinstance(raw_date, str):
                        if date_format == 'auto':
                            raw_date_clean = raw_date
                            if 'T' in raw_date:
                                if '+' in raw_date:
                                    raw_date_clean = raw_date.split('+')[0]
                                elif 'Z' in raw_date:
                                    raw_date_clean = raw_date.replace('Z', '')
                                raw_date_clean = raw_date_clean.replace('T', ' ')
                                try:
                                    dt = datetime.strptime(raw_date_clean, '%Y-%m-%d %H:%M:%S')
                                except ValueError:
                                    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d.%m.%Y %H:%M:%S', '%d.%m.%Y', '%m/%d/%Y %H:%M:%S', '%m/%d/%Y']:
                                        try:
                                            dt = datetime.strptime(raw_date, fmt)
                                            break
                                        except:
                                            continue
                                    else:
                                        errors.append(f"Не удалось распознать дату: {raw_date}")
                                        continue
                            else:
                                for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d.%m.%Y %H:%M:%S', '%d.%m.%Y', '%m/%d/%Y %H:%M:%S', '%m/%d/%Y']:
                                    try:
                                        dt = datetime.strptime(raw_date, fmt)
                                        break
                                    except:
                                        continue
                                else:
                                    errors.append(f"Не удалось распознать дату: {raw_date}")
                                    continue
                        else:
                            dt = datetime.strptime(raw_date, date_format)
                    else:
                        errors.append(f"Неподдерживаемый тип даты: {type(raw_date)}")
                        continue
                except Exception as e:
                    errors.append(f"Ошибка парсинга даты для {sn}: {raw_date} -> {e}")
                    continue

                notes = str(row.get(col_notes, '')).strip() if col_notes else ''

                obj, created = Reading.objects.update_or_create(
                    device=device,
                    timestamp=dt,
                    direction=direction,
                    defaults={'reading_value': val, 'notes': notes}
                )
                if created:
                    success += 1
                else:
                    updated += 1

        session_data = session.get_decoded()
        session_data['import_result'] = {
            'success': success,
            'updated': updated,
            'errors': errors[:10],
            'skipped': skipped[:10],
            'skipped_empty': skipped_empty[:10],
            'error_count': len(errors),
            'skipped_count': len(skipped),
            'empty_count': len(skipped_empty),
        }
        session_data.pop('import_data', None)
        session_data.pop('import_params', None)
        session.save()
        logger.info(f"Фоновый импорт завершён: добавлено {success}, обновлено {updated}")
    except Exception as e:
        logger.error(f"Ошибка в фоновом импорте: {e}")

@login_required
@user_passes_test(is_allowed_user)
def bulk_readings(request):
    # Очистка результата импорта
    if request.GET.get('clear_result'):
        request.session.pop('import_result', None)
        return redirect('bulk_readings')

    # Ручной ввод
    if request.method == 'POST' and 'manual_submit' in request.POST:
        date_str = request.POST.get('date')
        readings_text = request.POST.get('readings_text', '')
        notes = request.POST.get('notes', '')
        direction = request.POST.get('direction', 'aplus')
        if date_str:
            try:
                dt = datetime.strptime(date_str, "%Y-%m-%d")
            except:
                dt = datetime.now()
        else:
            dt = datetime.now()
        dt = dt.replace(hour=0, minute=0, second=0, microsecond=0)

        lines = readings_text.strip().split('\n')
        added = 0
        errors = []
        not_found = []
        for line in lines:
            parts = line.strip().split()
            if len(parts) < 2:
                errors.append(f"Неверный формат: {line}")
                continue
            sn = parts[0]
            val = parts[1]
            device = find_device(sn)
            if not device:
                not_found.append(sn)
                continue
            try:
                Reading.objects.update_or_create(
                    device=device,
                    timestamp=dt,
                    direction=direction,
                    defaults={'reading_value': float(val), 'notes': notes}
                )
                added += 1
            except Exception as e:
                errors.append(f"Ошибка для {sn}: {e}")
        messages.success(request, f"Добавлено показаний: {added}. Ошибок: {len(errors)}. Не найдено: {len(not_found)}")
        if errors:
            messages.warning(request, "Ошибки: " + "; ".join(errors[:3]))
        if not_found:
            messages.warning(request, "Не найдены: " + ", ".join(not_found[:3]))
        return redirect('bulk_readings')

    # Загрузка Excel (первый шаг – предпросмотр)
    if request.method == 'POST' and 'excel_submit' in request.POST and request.FILES.get('excel_file'):
        file = request.FILES['excel_file']
        try:
            df = pd.read_excel(file, engine='openpyxl')
        except Exception as e:
            messages.error(request, f'Ошибка чтения файла: {e}')
            return redirect('bulk_readings')

        if df.empty:
            messages.error(request, 'Файл пуст.')
            return redirect('bulk_readings')

        request.session['import_data'] = df.to_dict('records')
        request.session['import_columns'] = list(df.columns)
        request.session['import_shape'] = df.shape

        sample_data = df.head(10).values.tolist()
        return render(request, 'meters/import_readings_preview.html', {
            'columns': list(df.columns),
            'sample': sample_data,
            'total_rows': df.shape[0],
        })

    # Подтверждение импорта (запуск фона)
    if request.method == 'POST' and 'confirm_import' in request.POST:
        request.session['import_params'] = {
            'col_serial': request.POST.get('col_serial'),
            'col_timestamp': request.POST.get('col_timestamp'),
            'col_value': request.POST.get('col_value'),
            'col_notes': request.POST.get('col_notes'),
            'col_direction': request.POST.get('col_direction'),
            'default_direction': request.POST.get('default_direction', 'aplus'),
            'date_format': request.POST.get('date_format', 'auto'),
        }
        thread = threading.Thread(target=run_import_background, args=(request.session.session_key,))
        thread.daemon = True
        thread.start()
        messages.info(request, 'Импорт запущен в фоновом режиме. Обновите страницу через некоторое время для просмотра результатов.')
        return redirect('bulk_readings')

    return render(request, 'meters/bulk_readings.html')

# ===== МОДЕЛИ =====
@login_required
@user_passes_test(is_allowed_user)
def models(request):
    models_list = MeterModel.objects.all()
    context = {'models': models_list}
    return render(request, 'meters/models.html', context)

@login_required
@user_passes_test(is_allowed_user)
def add_model(request):
    if request.method == 'POST':
        try:
            model = MeterModel(
                catalog_code=request.POST.get('catalog_code'),
                model_name=request.POST.get('model_name'),
                digit_capacity=request.POST.get('digit_capacity'),
                phases=request.POST.get('phases'),
                nominal_current=request.POST.get('nominal_current'),
                nominal_voltage=request.POST.get('nominal_voltage'),
                system_type=request.POST.get('system_type'),
                period=request.POST.get('period'),
                device_type_id=request.POST.get('device_type_id'),
                device_type_str=request.POST.get('device_type_str'),
            )
            model.save()
            messages.success(request, f"Тип ПУ {model.model_name} добавлена.")
        except Exception as e:
            messages.error(request, f"Ошибка: {e}")
        return redirect('models')
    return render(request, 'meters/add_model.html')

@login_required
@user_passes_test(is_allowed_user)
def edit_model(request, model_id):
    model = get_object_or_404(MeterModel, id=model_id)
    if request.method == 'POST':
        try:
            model.catalog_code = request.POST.get('catalog_code')
            model.model_name = request.POST.get('model_name')
            model.digit_capacity = request.POST.get('digit_capacity')
            model.phases = request.POST.get('phases')
            model.nominal_current = request.POST.get('nominal_current')
            model.nominal_voltage = request.POST.get('nominal_voltage')
            model.system_type = request.POST.get('system_type')
            model.period = request.POST.get('period')
            model.device_type_id = request.POST.get('device_type_id')
            model.device_type_str = request.POST.get('device_type_str')
            model.save()
            messages.success(request, f"Тип ПУ {model.model_name} обновлена.")
        except Exception as e:
            messages.error(request, f"Ошибка: {e}")
        return redirect('models')
    context = {'model': model}
    return render(request, 'meters/edit_model.html', context)

@login_required
@user_passes_test(is_allowed_user)
def delete_model(request, model_id):
    model = get_object_or_404(MeterModel, id=model_id)
    if request.method == 'POST':
        model.delete()
        messages.success(request, f"Тип ПУ удалена.")
        return redirect('models')
    return render(request, 'meters/confirm_delete.html', {'object': model, 'type': 'Тип ПУ'})

def producer_stats_table(request):
    period = request.GET.get('period', 'today')
    region_id = request.GET.get('region_id')
    cache_key = f'producer_stats_table_{period}_{region_id}'
    html = cache.get(cache_key)
    if html is None:
        today = timezone.now().date()
        if period == 'today':
            start_date = today
        elif period == '3d':
            start_date = today - timedelta(days=3)
        elif period == '7d':
            start_date = today - timedelta(days=7)
        elif period == '30d':
            start_date = today - timedelta(days=30)
        else:
            start_date = today

        start_dt = datetime.combine(start_date, datetime.min.time())
        end_dt = datetime.combine(today, datetime.max.time())

        devices = Device.objects.filter(status='active').select_related('model')
        if region_id:
            devices = devices.filter(region_id=region_id)
        # devices = Device.objects.filter(status='active').select_related('model').only(
        #     'id', 'model__device_type_str', 'model__device_type_id'
        # )
        active_device_ids = set(
            Reading.objects.filter(
                device__status='active',
                timestamp__gte=start_dt,
                timestamp__lte=end_dt
            ).values_list('device_id', flat=True).distinct()
        )

        producer_stats_dict = {}
        for device in devices:
            model = device.model
            if model:
                code = model.device_type_str or ""
                type_id = model.device_type_id or ""
                display_name = get_producer_display(code, type_id)
            else:
                display_name = "Неизвестно"
            if display_name not in producer_stats_dict:
                producer_stats_dict[display_name] = {'total': 0, 'active': 0}
            producer_stats_dict[display_name]['total'] += 1
            if device.id in active_device_ids:
                producer_stats_dict[display_name]['active'] += 1

        producer_stats = []
        for name, stats in producer_stats_dict.items():
            percent = round((stats['active'] / stats['total'] * 100), 1) if stats['total'] else 0
            producer_stats.append({
                'producer': name,
                'total': stats['total'],
                'active': stats['active'],
                'percent': percent,
            })
        producer_stats.sort(key=lambda x: x['producer'])

        html = render_to_string('meters/_producer_table.html', {'producer_stats': producer_stats})
        cache.set(cache_key, html, 600)
    return HttpResponse(html)


# @login_required
# @user_passes_test(is_allowed_user)
def missing_readings_report(request):
    period = request.GET.get('period', 'today')
    specific_date = request.GET.get('date', '')          # НОВОЕ: конкретная дата
    producer_filter = request.GET.get('producer', '')
    model_filter = request.GET.get('model', '')
    status_filter = request.GET.get('status', '')
    region_filter = request.GET.get('region', '')
    group_filter = request.GET.get('group', '')
    export = request.GET.get('export', False)
    online_filter = request.GET.get('online', '')
    category_filter = request.GET.get('category', '')

    today = timezone.now().date()

    # Базовый запрос
    devices_qs = Device.objects.select_related('model')

    # --- Если указана конкретная дата ---
    if specific_date:
        try:
            target_date = datetime.strptime(specific_date, '%Y-%m-%d').date()
        except ValueError:
            target_date = today

        # Аннотация: есть ли показание на эту дату
        reading_exists = Exists(
            Reading.objects.filter(
                device=OuterRef('pk'),
                timestamp__date=target_date
            )
        )
        devices_qs = devices_qs.annotate(
            has_reading_on_date=reading_exists
        ).filter(has_reading_on_date=False)

        # Для отображения последнего показания (можно оставить как есть)
        last_reading_subquery = Reading.objects.filter(
            device=OuterRef('pk')
        ).order_by('-timestamp').values('timestamp')[:1]
        devices_qs = devices_qs.annotate(
            last_reading_date=Subquery(last_reading_subquery)
        )
        # Статус связи (последние 7 дней)
        seven_days_ago = timezone.now() - timedelta(days=7)
        devices_qs = devices_qs.annotate(
            has_recent_reading=Exists(
                Reading.objects.filter(
                    device=OuterRef('pk'),
                    timestamp__gte=seven_days_ago
                )
            )
        )
        # Для пагинации и шаблона будем использовать start_date = target_date (для совместимости)
        start_date = target_date
    else:
        # --- Старый режим: период ---
        if period == 'today':
            start_date = today
        elif period == '3d':
            start_date = today - timedelta(days=3)
        elif period == '7d':
            start_date = today - timedelta(days=7)
        elif period == '30d':
            start_date = today - timedelta(days=30)
        else:
            start_date = today

        last_reading_subquery = Reading.objects.filter(
            device=OuterRef('pk')
        ).order_by('-timestamp').values('timestamp')[:1]
        devices_qs = devices_qs.annotate(
            last_reading_date=Subquery(last_reading_subquery)
        )

        seven_days_ago = timezone.now() - timedelta(days=7)
        devices_qs = devices_qs.annotate(
            has_recent_reading=Exists(
                Reading.objects.filter(
                    device=OuterRef('pk'),
                    timestamp__gte=seven_days_ago
                )
            )
        )

        # Подсчёт количества показаний за период
        readings_subquery = Reading.objects.filter(
            device_id=OuterRef('id'),
            timestamp__date__gte=start_date,
            timestamp__date__lte=today
        ).values('device_id').annotate(cnt=Count('id')).values('cnt')

        devices_qs = devices_qs.annotate(
            readings_count=Coalesce(readings_subquery, Value(0))
        ).filter(readings_count=0)

    # --- Применяем остальные фильтры (статус, Тип ПУ, производитель, регион, группа) ---
    if status_filter:
        devices_qs = devices_qs.filter(status=status_filter)
    if model_filter:
        devices_qs = devices_qs.filter(model__id=model_filter)
    if producer_filter:
        devices_qs = devices_qs.filter(model__device_type_str=producer_filter)
    if region_filter:
        devices_qs = devices_qs.filter(region_id=region_filter)
    if group_filter:
        devices_qs = devices_qs.filter(groups__id=group_filter)
    if category_filter == '__none__':
        devices_qs = devices_qs.filter(category__isnull=True)
    elif category_filter:
        devices_qs = devices_qs.filter(category=category_filter)
    if online_filter == 'yes':
        devices_qs = devices_qs.filter(has_recent_reading=True)
    elif online_filter == 'no':
        devices_qs = devices_qs.filter(has_recent_reading=False)
    

    # --- Экспорт ---
    if export == '1':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="missing_readings_{specific_date or period}_{today}.csv"'
        writer = csv.writer(response)
        writer.writerow([
            'Заводской номер', 'Тип ПУ', 'Статус реестра',
            'Статус связи', 'Производитель', 'Ток', 'Фазность', 'Категория', 'Последнее показание'
        ])
        for device in devices_qs:
            producer = device.model.device_type_str if device.model else ''
            writer.writerow([
                device.serial_number,
                device.model.model_name if device.model else '',
                device.status,
                'Онлайн' if device.has_recent_reading else 'Оффлайн',
                producer,
                device.nominal_current,
                device.model.phases if device.model else '',
                device.get_category_display() or '',
                device.last_reading_date.strftime('%Y-%m-%d %H:%M:%S') if device.last_reading_date else '',
            ])
        return response

    if export == 'xlsx':
        import openpyxl
        from openpyxl.styles import Font, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Счетчики без показаний"
        headers = [
            'Заводской номер', 'Тип ПУ', 'Статус реестра',
            'Статус связи', 'Производитель', 'Ток', 'Фазность', 'Категория', 'Последнее показание'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        for row_idx, device in enumerate(devices_qs, 2):
            producer = device.model.device_type_str if device.model else ''
            ws.cell(row=row_idx, column=1, value=device.serial_number)
            ws.cell(row=row_idx, column=2, value=device.model.model_name if device.model else '')
            ws.cell(row=row_idx, column=3, value=device.status)
            ws.cell(row=row_idx, column=4, value='Онлайн' if device.has_recent_reading else 'Оффлайн')
            ws.cell(row=row_idx, column=5, value=producer)
            ws.cell(row=row_idx, column=6, value=device.nominal_current)
            ws.cell(row=row_idx, column=7, value=device.model.phases if device.model else '')
            ws.cell(row=row_idx, column=8, value=device.get_category_display() or '')
            ws.cell(row=row_idx, column=9, value=device.last_reading_date.strftime('%Y-%m-%d %H:%M:%S') if device.last_reading_date else '')
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="missing_readings_{specific_date or period}_{today}.xlsx"'
        wb.save(response)
        return response

    # --- Пагинация ---
    page = int(request.GET.get('page', 1))
    per_page = 10
    total = devices_qs.count()
    total_pages = (total + per_page - 1) // per_page if total > 0 else 1
    if page > total_pages:
        page = total_pages
    devices_list = devices_qs.order_by('serial_number')[(page-1)*per_page:page*per_page]
    page_range = range(1, total_pages + 1)

    # Добавляем producer_display для отображения
    for device in devices_list:
        if device.model:
            code = device.model.device_type_str or ""
            type_id = device.model.device_type_id or ""
            device.producer_display = get_producer_display(code, type_id)
        else:
            device.producer_display = "Неизвестно"

    # Подготовка данных для select-ов
    periods = [
        ('today', 'Сегодня', period == 'today'),
        ('3d', '3 дня', period == '3d'),
        ('7d', '7 дней', period == '7d'),
        ('30d', '30 дней', period == '30d'),
    ]

    codes = MeterModel.objects.values_list('device_type_str', flat=True).distinct().order_by('device_type_str')
    producers_with_selected = []
    for code in codes:
        display = get_producer_display(code, '')
        if display == "Неизвестно" and code:
            display = code
        selected = (code == producer_filter)
        producers_with_selected.append((display, code, selected))

    models_list = MeterModel.objects.all().order_by('model_name')
    models_with_selected = [(m, str(m.id) == model_filter) for m in models_list]

    statuses_with_selected = [(s, s == status_filter) for s in ['active', 'repair', 'inactive', 'temporary_off']]

    regions = Region.objects.all().order_by('name')
    regions_with_selected = [(r, str(r.id) == region_filter) for r in regions]
    groups = DeviceGroup.objects.all().order_by('name')
    groups_with_selected = [(g, str(g.id) == group_filter) for g in groups]

    categories = list(Device.CATEGORY_CHOICES)
    categories_with_selected = [(c[0], c[1], c[0] == category_filter) for c in categories]
    categories_with_selected.insert(0, ('', 'Все', category_filter == ''))
    categories_with_selected.append(('__none__', 'Не указано', category_filter == '__none__'))

    context = {
        'devices': devices_list,
        'total': total,
        'page': page,
        'total_pages': total_pages,
        'page_range': page_range,
        'period': period,
        'specific_date': specific_date,          # передаём в шаблон
        'producer_filter': producer_filter,
        'model_filter': model_filter,
        'status_filter': status_filter,
        'start_index': (page - 1) * per_page,
        'periods': periods,
        'producers_with_selected': producers_with_selected,
        'models_with_selected': models_with_selected,
        'statuses_with_selected': statuses_with_selected,
        'start_date': start_date,
        'regions_with_selected': regions_with_selected,
        'groups_with_selected': groups_with_selected,
        'selected_region': region_filter,
        'selected_group': group_filter,
        'categories_with_selected': categories_with_selected,
    }

    current_params = request.GET.copy()
    if 'page' in current_params:
        current_params.pop('page')
    current_params = current_params.urlencode()
    context['current_params'] = current_params

    return render(request, 'meters/missing_readings.html', context)

# исправление статуса active, inactive
@login_required
@user_passes_test(is_allowed_user)
def bulk_update_status_page(request):
    if request.method == 'POST':
        serials_text = request.POST.get('serials', '').strip()
        new_status = request.POST.get('status', 'active')
        if not serials_text:
            messages.error(request, 'Список серийных номеров пуст.')
            return redirect('bulk_update_status_page')
        serials = [s.strip() for s in serials_text.split('\n') if s.strip()]
        updated = 0
        not_found = []
        for sn in serials:
            device = Device.objects.filter(serial_number=sn).first()
            if device:
                device.status = new_status
                device.save()
                updated += 1
            else:
                not_found.append(sn)
        if updated:
            messages.success(request, f'Обновлено устройств: {updated}')
        if not_found:
            messages.warning(request, f'Не найдены: {", ".join(not_found[:10])}')
        return redirect('bulk_update_status_page')
    return render(request, 'meters/bulk_update_status.html')

# импорт показания Exsel
@login_required
@user_passes_test(is_allowed_user)
def import_readings_excel(request):
    import logging
    logger = logging.getLogger(__name__)

    if request.method == 'POST' and request.FILES.get('excel_file'):
        file = request.FILES['excel_file']
        try:
            import pandas as pd
            df = pd.read_excel(file, engine='openpyxl')
        except Exception as e:
            messages.error(request, f'Ошибка чтения файла: {e}')
            return redirect('import_readings_excel')

        if df.empty:
            messages.error(request, 'Файл пуст.')
            return redirect('import_readings_excel')

        # Сохраняем данные в сессию
        request.session['import_data'] = df.to_dict('records')
        request.session['import_columns'] = list(df.columns)

        sample_data = df.head(10).values.tolist()
        return render(request, 'meters/import_readings_preview.html', {
            'columns': list(df.columns),
            'sample': sample_data,
            'total_rows': df.shape[0],
        })

    if request.method == 'POST' and 'confirm' in request.POST:
        col_serial = request.POST.get('col_serial')
        col_timestamp = request.POST.get('col_timestamp')
        col_value = request.POST.get('col_value')
        col_notes = request.POST.get('col_notes')
        date_format = request.POST.get('date_format', 'auto')

        if not all([col_serial, col_timestamp, col_value]):
            messages.error(request, 'Необходимо указать колонки для серийного номера, даты и значения.')
            return redirect('import_readings_excel')

        data = request.session.get('import_data', [])
        if not data:
            messages.error(request, 'Данные не найдены, попробуйте загрузить файл заново.')
            return redirect('import_readings_excel')

        success = 0
        errors = []
        skipped = []
        updated = 0
        import pandas as pd
        from datetime import datetime

        logger.info(f"Начинаем импорт {len(data)} строк из Excel")

        for idx, row in enumerate(data):
            sn = str(row.get(col_serial, '')).strip()
            if not sn:
                errors.append(f"Строка {idx+1}: пустой серийный номер")
                continue

            device = find_device(sn)
            if not device:
                skipped.append(f"{sn} (строка {idx+1})")
                continue

            raw_date = row.get(col_timestamp)
            if pd.isna(raw_date):
                errors.append(f"Строка {idx+1} ({sn}): пустая дата")
                continue

            try:
                if isinstance(raw_date, datetime):
                    dt = raw_date
                elif isinstance(raw_date, str):
                    if date_format == 'auto':
                        # Пробуем форматы
                        if 'T' in raw_date:
                            if '+' in raw_date:
                                raw_date_clean = raw_date.split('+')[0]
                            elif 'Z' in raw_date:
                                raw_date_clean = raw_date.replace('Z', '')
                            else:
                                raw_date_clean = raw_date
                            for fmt in ['%Y-%m-%dT%H:%M:%S', '%Y-%m-%dT%H:%M', '%Y-%m-%d']:
                                try:
                                    dt = datetime.strptime(raw_date_clean, fmt)
                                    break
                                except:
                                    continue
                            else:
                                for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d.%m.%Y %H:%M:%S', '%d.%m.%Y', '%m/%d/%Y %H:%M:%S', '%m/%d/%Y']:
                                    try:
                                        dt = datetime.strptime(raw_date, fmt)
                                        break
                                    except:
                                        continue
                                else:
                                    errors.append(f"Строка {idx+1} ({sn}): не удалось распознать дату '{raw_date}'")
                                    continue
                        else:
                            for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d.%m.%Y %H:%M:%S', '%d.%m.%Y', '%m/%d/%Y %H:%M:%S', '%m/%d/%Y']:
                                try:
                                    dt = datetime.strptime(raw_date, fmt)
                                    break
                                except:
                                    continue
                            else:
                                errors.append(f"Строка {idx+1} ({sn}): не удалось распознать дату '{raw_date}'")
                                continue
                    else:
                        dt = datetime.strptime(raw_date, date_format)
                else:
                    errors.append(f"Строка {idx+1} ({sn}): неподдерживаемый тип даты {type(raw_date)}")
                    continue
            except Exception as e:
                errors.append(f"Строка {idx+1} ({sn}): ошибка парсинга даты '{raw_date}': {e}")
                continue

            try:
                val = float(row.get(col_value))
            except Exception as e:
                errors.append(f"Строка {idx+1} ({sn}): ошибка преобразования значения '{row.get(col_value)}': {e}")
                continue

            notes = str(row.get(col_notes, '')).strip() if col_notes else ''

            obj, created = Reading.objects.update_or_create(
                device=device,
                timestamp=dt,
                direction=direction,
                defaults={'reading_value': val, 'notes': notes}
            )
            if created:
                success += 1
            else:
                updated += 1

        logger.info(f"Импорт завершён: добавлено {success}, обновлено {updated}, ошибок {len(errors)}, пропущено {len(skipped)}")

        messages.success(request, f'✅ Импорт завершён. Добавлено: {success}, обновлено: {updated}.')
        if errors:
            messages.warning(request, f'⚠️ Ошибок при импорте: {len(errors)} (первые 5 показаны).')
            # Покажем первые 5 ошибок в отдельном сообщении
            for err in errors[:5]:
                messages.warning(request, f'  • {err}')
        if skipped:
            messages.warning(request, f'⚠️ Пропущено (устройство не найдено): {len(skipped)} (первые 5 показаны).')
            for sk in skipped[:5]:
                messages.warning(request, f'  • {sk}')

        request.session.pop('import_data', None)
        request.session.pop('import_columns', None)
        return redirect('bulk_readings')  # или 'devices', но лучше на страницу импорта

    return render(request, 'meters/import_readings_excel.html')

# ===== ОТЧЁТ БАЛАНСА =====
from decimal import Decimal

@login_required
@user_passes_test(is_allowed_user)
def balance_report(request):
    from django.db import connection
    from datetime import datetime

    date_str = request.GET.get('date', datetime.now().strftime('%Y-%m-%d'))
    level = request.GET.get('level', 'tp')
    export = request.GET.get('export', '')

    try:
        report_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        report_date = datetime.now().date()
        date_str = report_date.strftime('%Y-%m-%d')

    # Формируем запрос в зависимости от уровня
    if level == 'tp':
        query = """
            SELECT
                t.id AS object_id,
                t.name AS object_name,
                f.name AS parent_name,
                r.name AS region_name,
                hd.serial_number AS head_serial,
                hr.reading_value AS head_reading,
                COALESCE(SUM(cr.reading_value), 0) AS sum_children,
                (COALESCE(hr.reading_value, 0) - COALESCE(SUM(cr.reading_value), 0)) AS balance
            FROM transformer_substations t
            JOIN feeders f ON f.id = t.feeder_id
            JOIN substations s ON s.id = f.substation_id
            JOIN regions r ON r.id = s.region_id
            LEFT JOIN devices hd ON hd.id = t.head_meter_id
            LEFT JOIN readings hr ON hr.device_id = hd.id AND DATE(hr.timestamp) = %s
            LEFT JOIN devices cd ON cd.tp_id = t.id AND cd.id != t.head_meter_id
            LEFT JOIN readings cr ON cr.device_id = cd.id AND DATE(cr.timestamp) = %s
            GROUP BY t.id, t.name, f.name, r.name, hd.serial_number, hr.reading_value
            ORDER BY t.name
        """
        params = [report_date, report_date]
        columns = ['ТП', 'Фидер', 'Район', 'Головной счётчик', 'Показание', 'Сумма подчинённых', 'Баланс']
        header = ['object_name', 'parent_name', 'region_name', 'head_serial', 'head_reading', 'sum_children', 'balance']
    elif level == 'feeder':
        query = """
            SELECT
                f.id AS object_id,
                f.name AS object_name,
                s.name AS parent_name,
                r.name AS region_name,
                hd.serial_number AS head_serial,
                hr.reading_value AS head_reading,
                COALESCE(SUM(cr.reading_value), 0) AS sum_children,
                (COALESCE(hr.reading_value, 0) - COALESCE(SUM(cr.reading_value), 0)) AS balance
            FROM feeders f
            JOIN substations s ON s.id = f.substation_id
            JOIN regions r ON r.id = s.region_id
            LEFT JOIN devices hd ON hd.id = f.head_meter_id
            LEFT JOIN readings hr ON hr.device_id = hd.id AND DATE(hr.timestamp) = %s
            LEFT JOIN devices cd ON cd.feeder_id = f.id AND cd.id != f.head_meter_id
            LEFT JOIN readings cr ON cr.device_id = cd.id AND DATE(cr.timestamp) = %s
            GROUP BY f.id, f.name, s.name, r.name, hd.serial_number, hr.reading_value
            ORDER BY f.name
        """
        params = [report_date, report_date]
        columns = ['Фидер', 'Подстанция', 'Район', 'Головной счётчик', 'Показание', 'Сумма подчинённых', 'Баланс']
        header = ['object_name', 'parent_name', 'region_name', 'head_serial', 'head_reading', 'sum_children', 'balance']
    elif level == 'substation':
        query = """
            SELECT
                s.id AS object_id,
                s.name AS object_name,
                r.name AS parent_name,
                '' AS region_name,
                hd.serial_number AS head_serial,
                hr.reading_value AS head_reading,
                COALESCE(SUM(cr.reading_value), 0) AS sum_children,
                (COALESCE(hr.reading_value, 0) - COALESCE(SUM(cr.reading_value), 0)) AS balance
            FROM substations s
            JOIN regions r ON r.id = s.region_id
            LEFT JOIN devices hd ON hd.id = s.head_meter_id
            LEFT JOIN readings hr ON hr.device_id = hd.id AND DATE(hr.timestamp) = %s
            LEFT JOIN devices cd ON cd.substation_id = s.id AND cd.id != s.head_meter_id
            LEFT JOIN readings cr ON cr.device_id = cd.id AND DATE(cr.timestamp) = %s
            GROUP BY s.id, s.name, r.name, hd.serial_number, hr.reading_value
            ORDER BY s.name
        """
        params = [report_date, report_date]
        columns = ['Подстанция', 'Район', 'Головной счётчик', 'Показание', 'Сумма подчинённых', 'Баланс']
        header = ['object_name', 'parent_name', 'head_serial', 'head_reading', 'sum_children', 'balance']
    else:  # region
        query = """
            SELECT
                r.id AS object_id,
                r.name AS object_name,
                '' AS parent_name,
                '' AS region_name,
                hd.serial_number AS head_serial,
                hr.reading_value AS head_reading,
                COALESCE(SUM(cr.reading_value), 0) AS sum_children,
                (COALESCE(hr.reading_value, 0) - COALESCE(SUM(cr.reading_value), 0)) AS balance
            FROM regions r
            LEFT JOIN devices hd ON hd.id = r.head_meter_id
            LEFT JOIN readings hr ON hr.device_id = hd.id AND DATE(hr.timestamp) = %s
            LEFT JOIN devices cd ON cd.region_id = r.id AND cd.id != r.head_meter_id
            LEFT JOIN readings cr ON cr.device_id = cd.id AND DATE(cr.timestamp) = %s
            GROUP BY r.id, r.name, hd.serial_number, hr.reading_value
            ORDER BY r.name
        """
        params = [report_date, report_date]
        columns = ['Район', 'Головной счётчик', 'Показание', 'Сумма подчинённых', 'Баланс']
        header = ['object_name', 'head_serial', 'head_reading', 'sum_children', 'balance']

    with connection.cursor() as cursor:
        cursor.execute(query, params)
        rows = cursor.fetchall()

    data = []
    for row in rows:
        item = {}
        for i, col in enumerate(header):
            val = row[i]
            if isinstance(val, Decimal):
                val = float(val)
            item[col] = val if val is not None else 0
        data.append(item)

    # Экспорт
    if export == 'excel':
        import openpyxl
        from openpyxl.styles import Font, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Баланс {report_date}"
        for col, title in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=title)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        for row_idx, item in enumerate(data, 2):
            for col_idx, key in enumerate(header, 1):
                ws.cell(row=row_idx, column=col_idx, value=item.get(key, ''))
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="balance_{report_date}_{level}.xlsx"'
        wb.save(response)
        return response

    if export == 'csv':
        import csv
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="balance_{report_date}_{level}.csv"'
        writer = csv.writer(response)
        writer.writerow(columns)
        for item in data:
            row = [item.get(key, '') for key in header]
            writer.writerow(row)
        return response

    levels = [
        ('tp', 'ТП', level == 'tp'),
        ('feeder', 'Фидер', level == 'feeder'),
        ('substation', 'Подстанция', level == 'substation'),
        ('region', 'Район', level == 'region'),
    ]

    context = {
        'data': data,
        'columns': columns,
        'header': header,
        'date': date_str,
        'level': level,
        'levels': levels,
        'report_date': report_date,
    }
    return render(request, 'meters/balance.html', context)

import subprocess
import threading
import logging
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.csrf import csrf_exempt

logger = logging.getLogger(__name__)

@login_required
@user_passes_test(lambda u: u.is_staff)
@csrf_exempt
def restart_robot(request, robot_name):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    # Маппинг: имя робота -> команда Django (контейнер больше не нужен)
    robot_config = {
        'RiseSun': 'sync_risesun',
        'SunRise': 'sync_sunrise',
        'cEnergo': 'sync_mssql',
        'Hexing_KUK': 'sync_hexing',
        'Hexing_POP': 'sync_hxn',
        'Sanxing_old': 'sync_sanxing',
        'Sanxing_new_100A': 'sync_website',
        'Sanxing_new_5A': 'sync_website_period1',
        'Star': 'sync_star',
    }

    command = robot_config.get(robot_name)
    if not command:
        return JsonResponse({'error': f'Unknown robot: {robot_name}'}, status=400)

    def run_command():
        try:
            # Запускаем команду напрямую в текущем контейнере
            subprocess.Popen(
                ['python', 'manage.py', command, '--verbose'],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                cwd='/app'  # рабочая директория проекта
            )
            logger.info(f"Manual restart of {robot_name} initiated (direct call)")
        except Exception as e:
            logger.error(f"Manual restart of {robot_name} failed: {e}")

    thread = threading.Thread(target=run_command)
    thread.daemon = True
    thread.start()

    return JsonResponse({'message': f'Robot {robot_name} restart initiated'})

# ===== ОТЧЁТ: СЧЁТЧИКИ С ПОКАЗАНИЯМИ (НОВЫЙ) =====
# @login_required
# @user_passes_test(is_allowed_user)
def readings_with_report(request):
    period = request.GET.get('period', 'today')
    specific_date = request.GET.get('date', '')          # новая дата
    producer_filter = request.GET.get('producer', '')
    model_filter = request.GET.get('model', '')
    status_filter = request.GET.get('status', '')
    region_filter = request.GET.get('region', '')
    group_filter = request.GET.get('group', '')
    export = request.GET.get('export', False)
    online_filter = request.GET.get('online', '')
    category_filter = request.GET.get('category', '')

    today = timezone.now().date()
    if period == 'today':
        start_date = today
    elif period == '3d':
        start_date = today - timedelta(days=3)
    elif period == '7d':
        start_date = today - timedelta(days=7)
    elif period == '30d':
        start_date = today - timedelta(days=30)
    else:
        start_date = today

    # Базовый запрос
    devices_qs = Device.objects.select_related('model')

    # Если указана конкретная дата – используем её, иначе работаем с периодом
    if specific_date:
        try:
            target_date = datetime.strptime(specific_date, '%Y-%m-%d').date()
        except ValueError:
            target_date = today
        # Аннотируем значение показания на эту дату (берём первое за день, можно последнее)
        reading_subquery = Reading.objects.filter(
            device=OuterRef('pk'),
            timestamp__date=target_date
        ).order_by('-timestamp').values('reading_value')[:1]
        devices_qs = devices_qs.annotate(
            reading_on_date=Subquery(reading_subquery)
        )
        # Фильтруем только те, у которых есть показание на эту дату
        devices_qs = devices_qs.filter(reading_on_date__isnull=False)
        # Для отображения в таблице будем использовать это значение
        show_reading_value = True
        # Для статуса связи используем последние 7 дней
        seven_days_ago = timezone.now() - timedelta(days=7)
        devices_qs = devices_qs.annotate(
            has_recent_reading=Exists(
                Reading.objects.filter(
                    device=OuterRef('pk'),
                    timestamp__gte=seven_days_ago
                )
            )
        )
        # Последнее показание – тоже на эту дату (или можно оставить last_reading_date)
        last_reading_subquery = Reading.objects.filter(
            device=OuterRef('pk'),
            timestamp__date=target_date
        ).order_by('-timestamp').values('timestamp')[:1]
        devices_qs = devices_qs.annotate(
            last_reading_date=Subquery(last_reading_subquery)
        )
    else:
        # Старый режим: период
        last_reading_subquery = Reading.objects.filter(
            device=OuterRef('pk')
        ).order_by('-timestamp').values('timestamp')[:1]
        devices_qs = devices_qs.annotate(
            last_reading_date=Subquery(last_reading_subquery)
        )
        seven_days_ago = timezone.now() - timedelta(days=7)
        devices_qs = devices_qs.annotate(
            has_recent_reading=Exists(
                Reading.objects.filter(
                    device=OuterRef('pk'),
                    timestamp__gte=seven_days_ago
                )
            )
        )
        # Подсчёт количества показаний за период
        readings_subquery = Reading.objects.filter(
            device_id=OuterRef('id'),
            timestamp__date__gte=start_date,
            timestamp__date__lte=today
        ).values('device_id').annotate(cnt=Count('id')).values('cnt')
        devices_qs = devices_qs.annotate(
            readings_count=Coalesce(readings_subquery, Value(0))
        ).filter(readings_count__gt=0)
        # Поле значения на дату не нужно
        show_reading_value = False

    # Применяем остальные фильтры
    if status_filter:
        devices_qs = devices_qs.filter(status=status_filter)
    if model_filter:
        devices_qs = devices_qs.filter(model__id=model_filter)
    if producer_filter:
        devices_qs = devices_qs.filter(model__device_type_str=producer_filter)
    if region_filter:
        devices_qs = devices_qs.filter(region_id=region_filter)
    if group_filter:
        devices_qs = devices_qs.filter(groups__id=group_filter)
    if category_filter == '__none__':
        devices_qs = devices_qs.filter(category__isnull=True)
    elif category_filter:
        devices_qs = devices_qs.filter(category=category_filter)
    if online_filter == 'yes':
        devices_qs = devices_qs.filter(has_recent_reading=True)
    elif online_filter == 'no':
        devices_qs = devices_qs.filter(has_recent_reading=False)

    # Экспорт (CSV/Excel) – аналогично missing_readings, но с дополнительной колонкой "Значение"
    if export == '1':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="readings_with_{period}_{today}.csv"'
        writer = csv.writer(response)
        headers = [
            'Заводской номер', 'Тип ПУ', 'Статус реестра',
            'Статус связи', 'Производитель', 'Ток', 'Фазность', 'Категория'
        ]
        if show_reading_value and specific_date:
            headers.append(f'Значение на {specific_date}')
        else:
            headers.append('Последнее показание (дата)')
        writer.writerow(headers)

        for device in devices_qs:
            producer = device.model.device_type_str if device.model else ''
            row = [
                device.serial_number,
                device.model.model_name if device.model else '',
                device.status,
                'Онлайн' if device.has_recent_reading else 'Оффлайн',
                producer,
                device.nominal_current,
                device.model.phases if device.model else '',
                device.get_category_display() or '',
            ]
            if show_reading_value and specific_date:
                row.append(str(device.reading_on_date) if device.reading_on_date is not None else '')
            else:
                row.append(device.last_reading_date.strftime('%Y-%m-%d %H:%M:%S') if device.last_reading_date else '')
            writer.writerow(row)
        return response

    if export == 'xlsx':
        import openpyxl
        from openpyxl.styles import Font, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Счетчики с показаниями"
        headers = [
            'Заводской номер', 'Тип ПУ', 'Статус реестра',
            'Статус связи', 'Производитель', 'Ток', 'Фазность', 'Категория'
        ]
        if show_reading_value and specific_date:
            headers.append(f'Значение на {specific_date}')
        else:
            headers.append('Последнее показание (дата)')
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        for row_idx, device in enumerate(devices_qs, 2):
            producer = device.model.device_type_str if device.model else ''
            ws.cell(row=row_idx, column=1, value=device.serial_number)
            ws.cell(row=row_idx, column=2, value=device.model.model_name if device.model else '')
            ws.cell(row=row_idx, column=3, value=device.status)
            ws.cell(row=row_idx, column=4, value='Онлайн' if device.has_recent_reading else 'Оффлайн')
            ws.cell(row=row_idx, column=5, value=producer)
            ws.cell(row=row_idx, column=6, value=device.nominal_current)
            ws.cell(row=row_idx, column=7, value=device.model.phases if device.model else '')
            ws.cell(row=row_idx, column=8, value=device.get_category_display() or '')
            if show_reading_value and specific_date:
                ws.cell(row=row_idx, column=8, value=float(device.reading_on_date) if device.reading_on_date is not None else '')
            else:
                ws.cell(row=row_idx, column=8, value=device.last_reading_date.strftime('%Y-%m-%d %H:%M:%S') if device.last_reading_date else '')
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="readings_with_{period}_{today}.xlsx"'
        wb.save(response)
        return response

    # Пагинация
    page = int(request.GET.get('page', 1))
    per_page = 10
    total = devices_qs.count()
    total_pages = (total + per_page - 1) // per_page if total > 0 else 1
    if page > total_pages:
        page = total_pages
    devices_list = devices_qs.order_by('serial_number')[(page-1)*per_page:page*per_page]
    page_range = range(1, total_pages + 1)

    for device in devices_list:
        if device.model:
            code = device.model.device_type_str or ""
            type_id = device.model.device_type_id or ""
            device.producer_display = get_producer_display(code, type_id)
        else:
            device.producer_display = "Неизвестно"

    periods = [
        ('today', 'Сегодня', period == 'today'),
        ('3d', '3 дня', period == '3d'),
        ('7d', '7 дней', period == '7d'),
        ('30d', '30 дней', period == '30d'),
    ]

    codes = MeterModel.objects.values_list('device_type_str', flat=True).distinct().order_by('device_type_str')
    producers_with_selected = []
    for code in codes:
        display = get_producer_display(code, '')
        if display == "Неизвестно" and code:
            display = code
        selected = (code == producer_filter)
        producers_with_selected.append((display, code, selected))

    models_list = MeterModel.objects.all().order_by('model_name')
    models_with_selected = [(m, str(m.id) == model_filter) for m in models_list]

    statuses_with_selected = [(s, s == status_filter) for s in ['active', 'repair', 'inactive', 'temporary_off']]

    regions = Region.objects.all().order_by('name')
    regions_with_selected = [(r, str(r.id) == region_filter) for r in regions]
    groups = DeviceGroup.objects.all().order_by('name')
    groups_with_selected = [(g, str(g.id) == group_filter) for g in groups]

    categories = list(Device.CATEGORY_CHOICES)
    categories_with_selected = [(c[0], c[1], c[0] == category_filter) for c in categories]
    categories_with_selected.insert(0, ('', 'Все', category_filter == ''))
    categories_with_selected.append(('__none__', 'Не указано', category_filter == '__none__'))

    context = {
        'devices': devices_list,
        'total': total,
        'page': page,
        'total_pages': total_pages,
        'page_range': page_range,
        'period': period,
        'specific_date': specific_date,
        'producer_filter': producer_filter,
        'model_filter': model_filter,
        'status_filter': status_filter,
        'start_index': (page - 1) * per_page,
        'periods': periods,
        'producers_with_selected': producers_with_selected,
        'models_with_selected': models_with_selected,
        'statuses_with_selected': statuses_with_selected,
        'start_date': start_date,
        'regions_with_selected': regions_with_selected,
        'groups_with_selected': groups_with_selected,
        'selected_region': region_filter,
        'selected_group': group_filter,
        'show_reading_value': show_reading_value,  # флаг для шаблона
        'categories_with_selected': categories_with_selected,
    }

    current_params = request.GET.copy()
    if 'page' in current_params:
        current_params.pop('page')
    current_params = current_params.urlencode()
    context['current_params'] = current_params

    return render(request, 'meters/readings_with.html', context)

# ===== УПРАВЛЕНИЕ ГРУППАМИ =====
@login_required
@user_passes_test(is_allowed_user)
def groups_list(request):
    groups = DeviceGroup.objects.all().order_by('name')
    return render(request, 'meters/groups_list.html', {'groups': groups})

@login_required
@user_passes_test(is_allowed_user)
def group_create(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        if not name:
            messages.error(request, 'Название группы обязательно.')
        elif DeviceGroup.objects.filter(name=name).exists():
            messages.error(request, 'Группа с таким названием уже существует.')
        else:
            DeviceGroup.objects.create(name=name, description=description)
            messages.success(request, f'Группа "{name}" создана.')
            return redirect('groups_list')
    return render(request, 'meters/group_form.html', {'action': 'create'})

@login_required
@user_passes_test(is_allowed_user)
def group_edit(request, group_id):
    group = get_object_or_404(DeviceGroup, id=group_id)
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        if not name:
            messages.error(request, 'Название группы обязательно.')
        elif DeviceGroup.objects.exclude(id=group_id).filter(name=name).exists():
            messages.error(request, 'Группа с таким названием уже существует.')
        else:
            group.name = name
            group.description = description
            group.save()
            messages.success(request, f'Группа "{name}" обновлена.')
            return redirect('groups_list')
    return render(request, 'meters/group_form.html', {'group': group, 'action': 'edit'})

@login_required
@user_passes_test(is_allowed_user)
def group_delete(request, group_id):
    group = get_object_or_404(DeviceGroup, id=group_id)
    if request.method == 'POST':
        group.delete()
        messages.success(request, f'Группа "{group.name}" удалена.')
        return redirect('groups_list')
    return render(request, 'meters/confirm_delete.html', {'object': group, 'type': 'группу'})

from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse
from .registration import register_devices_for_robot, ROBOT_CONFIGS

@login_required
@user_passes_test(is_allowed_user)
def register_devices(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    robot_name = request.POST.get('robot_name')
    limit = request.POST.get('limit')
    limit = int(limit) if limit and limit.isdigit() else None
    
    if not robot_name:
        return JsonResponse({'error': 'robot_name required'}, status=400)
    
    if robot_name not in ROBOT_CONFIGS:
        return JsonResponse({'error': f'No registration config for {robot_name}'}, status=400)
    
    try:
        count = register_devices_for_robot(robot_name, limit)
        return JsonResponse({'message': f'Registered {count} new devices for {robot_name}', 'count': count})
    except Exception as e:
        logger.error(f"Error registering devices for {robot_name}: {e}")
        return JsonResponse({'error': str(e)}, status=500)

def register_devices_page(request):
    robots = list(ROBOT_CONFIGS.keys())
    return render(request, 'meters/register_devices.html', {'robots': robots})

from django.utils import timezone
from datetime import datetime

@login_required
@user_passes_test(is_allowed_user)
def registered_devices_report(request):
    # Получаем параметры фильтра
    source_filter = request.GET.get('source', '')
    search = request.GET.get('search', '').strip()
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    export = request.GET.get('export', False)

    # Базовый запрос – устройства, зарегистрированные через автоматическую регистрацию
    # Предполагаем, что при регистрации мы заполняем поле api_id значением источника (например, 'SR', 'SX', 'HK')
    devices_qs = Device.objects.select_related('model').filter(
        Q(api_id__isnull=False) & ~Q(api_id='')
    ).order_by('-created_at')

    # Фильтр по источнику
    if source_filter:
        devices_qs = devices_qs.filter(api_id=source_filter)

    # Поиск по серийному номеру
    if search:
        devices_qs = devices_qs.filter(serial_number__icontains=search)

    # Фильтр по дате регистрации
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            devices_qs = devices_qs.filter(created_at__date__gte=date_from_obj)
        except ValueError:
            pass
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            devices_qs = devices_qs.filter(created_at__date__lte=date_to_obj)
        except ValueError:
            pass

    # Экспорт в Excel
    if export == 'xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Зарегистрированные счетчики"

        headers = ['Заводской номер', 'Тип ПУ', 'Источник', 'Дата регистрации', 'Статус']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        for row_idx, device in enumerate(devices_qs, 2):
            ws.cell(row=row_idx, column=1, value=device.serial_number)
            ws.cell(row=row_idx, column=2, value=device.model.model_name if device.model else '')
            ws.cell(row=row_idx, column=3, value=device.api_id or '—')
            ws.cell(row=row_idx, column=4, value=device.created_at.strftime('%Y-%m-%d %H:%M:%S') if device.created_at else '')
            ws.cell(row=row_idx, column=5, value=device.status)

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="registered_devices_{timezone.now().date()}.xlsx"'
        wb.save(response)
        return response

    # Пагинация
    page = int(request.GET.get('page', 1))
    per_page = 20
    total = devices_qs.count()
    total_pages = (total + per_page - 1) // per_page if total > 0 else 1
    if page > total_pages:
        page = total_pages
    devices = devices_qs[(page-1)*per_page:page*per_page]
    page_range = range(1, total_pages + 1)

    # Список источников (роботов) для фильтра с флагом selected
    sources = [
        ('SR', 'SunRise', source_filter == 'SR'),
        ('SX', 'Sanxing_old', source_filter == 'SX'),
        ('HK', 'Hexing KUK', source_filter == 'HK'),
        # Добавьте другие источники по мере необходимости
    ]

    context = {
        'devices': devices,
        'total': total,
        'page': page,
        'total_pages': total_pages,
        'page_range': page_range,
        'sources': sources,
        'selected_source': source_filter,
        'search': search,
        'date_from': date_from,
        'date_to': date_to,
        'start_index': (page - 1) * per_page,
    }
    return render(request, 'meters/registered_devices.html', context)

@login_required
@user_passes_test(is_allowed_user)
def bulk_update_region_page(request):
    if request.method == 'POST':
        serials_text = request.POST.get('serials', '').strip()
        region_id = request.POST.get('region_id', '')
        substation_id = request.POST.get('substation_id', '')
        feeder_id = request.POST.get('feeder_id', '')
        tp_id = request.POST.get('tp_id', '')

        if not serials_text:
            messages.error(request, 'Список серийных номеров пуст.')
            return redirect('bulk_update_region_page')

        serials = [s.strip() for s in serials_text.split('\n') if s.strip()]
        if not serials:
            messages.error(request, 'Нет валидных номеров.')
            return redirect('bulk_update_region_page')

        # Подготавливаем поля для обновления
        update_fields = {}
        if region_id:
            update_fields['region_id'] = region_id
        if substation_id:
            update_fields['substation_id'] = substation_id
        if feeder_id:
            update_fields['feeder_id'] = feeder_id
        if tp_id:
            update_fields['tp_id'] = tp_id

        if not update_fields:
            messages.warning(request, 'Выберите хотя бы одно поле для обновления.')
            return redirect('bulk_update_region_page')

        updated = 0
        not_found = []
        for sn in serials:
            device = Device.objects.filter(serial_number=sn).first()
            if device:
                Device.objects.filter(id=device.id).update(**update_fields)
                updated += 1
            else:
                not_found.append(sn)

        if updated:
            messages.success(request, f'Обновлено устройств: {updated}')
        if not_found:
            messages.warning(request, f'Не найдены: {", ".join(not_found[:10])}')

        return redirect('bulk_update_region_page')

    # GET – показываем форму
    regions = Region.objects.all().order_by('name')
    substations = Substation.objects.all().order_by('name')
    feeders = Feeder.objects.all().order_by('name')
    tps = TransformerSubstation.objects.all().order_by('name')

    context = {
        'regions': regions,
        'substations': substations,
        'feeders': feeders,
        'tps': tps,
    }
    return render(request, 'meters/bulk_update_region.html', context)

# @login_required
# @user_passes_test(is_allowed_user)
def region_stats_table(request):
    period = request.GET.get('period', 'today')
    cache_key = f'region_stats_table_{period}'
    html = cache.get(cache_key)
    if html is None:
        today = timezone.now().date()
        if period == 'today':
            start_date = today
        elif period == '3d':
            start_date = today - timedelta(days=3)
        elif period == '7d':
            start_date = today - timedelta(days=7)
        elif period == '30d':
            start_date = today - timedelta(days=30)
        else:
            start_date = today

        start_dt = datetime.combine(start_date, datetime.min.time())
        end_dt = datetime.combine(today, datetime.max.time())

        # Получаем все активные устройства с районом
        devices = Device.objects.filter(status='active').select_related('region')

        # Аннотируем: есть ли показание за период
        from django.db.models import Exists, OuterRef, Count, Case, When, IntegerField, Value, Q
        readings_subquery = Reading.objects.filter(
            device=OuterRef('pk'),
            timestamp__gte=start_dt,
            timestamp__lte=end_dt
        ).values('device_id').annotate(cnt=Count('id')).values('cnt')
        devices = devices.annotate(
            has_reading=Exists(readings_subquery)
        )

        # Группируем по районам
        region_stats = devices.values('region__id', 'region__name').annotate(
            total=Count('id'),
            active=Count(Case(When(has_reading=True, then=1), output_field=IntegerField()))
        ).order_by('region__name')

        # Преобразуем для шаблона
        stats = []
        for stat in region_stats:
            total = stat['total']
            active = stat['active']
            offline = total - active
            percent = round((active / total * 100), 1) if total else 0
            stats.append({
                'region_id': stat['region__id'],
                'region_name': stat['region__name'] if stat['region__name'] else 'Без района',
                'total': total,
                'active': active,
                'offline': offline,
                'percent': percent,
            })

        html = render_to_string('meters/_region_table.html', {'region_stats': stats})
        cache.set(cache_key, html, 600)
    return HttpResponse(html)

# ===== СВЕРКА С 1С =====
def find_header_row(df, keywords=None):
    """Определяет строку с заголовками по наличию ключевых слов."""
    if keywords is None:
        keywords = ['заводской', 'зав', '№', 'счетчик', 'Тип ПУ', 'тип', 'лицевой']
    for idx, row in df.iterrows():
        # Преобразуем все ячейки строки в строки и объединяем
        row_text = ' '.join([str(cell).lower() for cell in row if pd.notna(cell)])
        if any(kw in row_text for kw in keywords):
            return idx
    return 0  # если не нашли – используем первую строку



# @login_required
def compare_1c(request):
    """
    Загрузка Excel-файла из 1С.
    - Обязательный выбор района из выпадающего списка.
    - Все проблемные заявки относятся к выбранному району.
    - В БД сохраняются только проблемные записи (не OK и не partial).
    """
    if request.GET.get('clear') == '1':
        request.session.pop('compare_results', None)
        request.session.pop('compare_total', None)
        messages.info(request, 'Данные сессии очищены.')
        return redirect('compare_1c')

    if request.method == 'POST' and request.FILES.get('excel_file'):
        file = request.FILES['excel_file']
        
        # --- Получаем выбранный район из формы ---
        region_id = request.POST.get('region_id')
        if not region_id:
            messages.error(request, 'Пожалуйста, выберите район для загрузки.')
            return redirect('compare_1c')
        
        try:
            default_region = Region.objects.get(id=region_id)
        except Region.DoesNotExist:
            messages.error(request, 'Выбранный район не найден.')
            return redirect('compare_1c')
        
        messages.info(request, f'Загрузка для района: "{default_region.name}"')

        try:
            df_raw = pd.read_excel(file, engine='openpyxl', header=None, dtype=str)
        except Exception as e:
            messages.error(request, f'Ошибка чтения файла: {e}')
            return redirect('compare_1c')

        if df_raw.empty:
            messages.error(request, 'Файл пуст.')
            return redirect('compare_1c')

        header_row = find_header_row(df_raw)
        df = pd.read_excel(file, engine='openpyxl', header=header_row, dtype=str)

        # Определяем колонки
        serial_col = None
        model_col = None
        for col in df.columns:
            col_lower = str(col).lower()
            if 'заводской' in col_lower or 'зав' in col_lower or '№' in col_lower:
                serial_col = col
            if 'счетчик' in col_lower or 'Тип ПУ' in col_lower or 'тип' in col_lower:
                model_col = col

        if serial_col is None:
            messages.error(request, 'Не найдена колонка с заводскими номерами.')
            return redirect('compare_1c')

        # --- Индексы устройств ---
        devices = Device.objects.filter(status__in=['active', 'temporary_off']).select_related('model', 'region')
        devices_by_norm = {}
        # devices_by_suffix = {}

        def norm_serial(s):
            if not s:
                return ''
            digits = re.sub(r'\D', '', str(s))
            return digits.lstrip('0') or '0'

        for d in devices:
            norm = norm_serial(d.serial_number)
            devices_by_norm[norm] = d
            # for i in range(8, min(len(norm), 16)):
            #     suffix = norm[-i:]
            #     if suffix not in devices_by_suffix:
            #         devices_by_suffix[suffix] = d

        today = timezone.now().date()
        created_count = 0
        updated_count = 0
        skipped_ok = 0
        skipped_partial = 0

        for idx, row in df.iterrows():
            serial_val = row[serial_col]
            if pd.isna(serial_val) or str(serial_val).strip() == '':
                continue

            # Очистка номера от .0
            serial_1c_raw = str(serial_val).strip()
            serial_1c_raw = re.sub(r'\.0+$', '', serial_1c_raw)
            serial_1c = serial_1c_raw

            model_1c = str(row[model_col]).strip() if model_col and not pd.isna(row[model_col]) else ''

            # Поиск устройства
            device = None
            norm = norm_serial(serial_1c)
            device = devices_by_norm.get(norm)
            # if not device and len(norm) >= 8:
            #     suffix = norm[-8:]
            #     device = devices_by_suffix.get(suffix)
            # if not device:
            #     device = devices_by_suffix.get(norm)

            # --- Определяем тип заявки ---
            found = device is not None
            model_db = ''
            model_match_type = 'none'
            is_online = False
            last_reading_value = None
            last_reading_date = None
            action_type = 'not_found'
            action_text = ''

            if device:
                model_db = device.model.model_name if device.model else ''
                model_match_type = compare_models(model_1c, model_db)

                last_reading = Reading.objects.filter(device=device).order_by('-timestamp').first()
                if last_reading:
                    last_reading_value = float(last_reading.reading_value)
                    last_reading_date = last_reading.timestamp
                    days = (timezone.now() - last_reading.timestamp).days
                    is_online = days < 7

                # --- ПРОВЕРКА НА ВРЕМЕННО ОТКЛЮЧЕННЫЙ СЧЕТЧИК (первая) ---
                if device.status == 'temporary_off':
                    action_type = 'temporary_off'
                    action_text = 'Счётчик временно отключен...'
                    # Сохраняем заявку и переходим к следующей строке
                    # (код сохранения будет ниже, но мы устанавливаем нужные переменные)
                    # Остальные проверки пропускаем
                else:
                    # --- Остальная логика для активных устройств ---
                    if model_match_type == 'exact' and is_online:
                        skipped_ok += 1
                        continue
                    elif model_match_type == 'partial':
                        skipped_partial += 1
                        continue
                    elif model_match_type in ('partial', 'none'):
                        action_type = 'fix_model'
                        action_text = f'Проверьте и исправьте Тип ПУ в 1С на «{model_db}».'
                    elif not is_online:
                        action_type = 'offline'
                        action_text = 'Счётчик оффлайн. Проверьте связь по инструкции.'
                    else:
                        skipped_ok += 1
                        continue
            else:
                # Не найден
                action_type = 'not_found'
                action_text = f'Не найден в базе. Проверить фактический ПУ'

            # --- Сохраняем только проблемные заявки с выбранным районом ---
            # Ищем существующую заявку по serial_1c (без учёта check_date)
            existing_issue = CompareResult.objects.filter(serial_1c=serial_1c).order_by('-check_date').first()

            if existing_issue:
                # Если заявка закрыта – создаём новую
                if existing_issue.status in ['fixed', 'auto_fixed', 'ignored']:
                    CompareResult.objects.create(
                        serial_1c=serial_1c,
                        model_1c=model_1c,
                        device=device,
                        model_db=model_db,
                        model_match_type=model_match_type,
                        is_online=is_online,
                        last_reading_date=last_reading_date,
                        last_reading_value=last_reading_value,
                        region=default_region,
                        action_type=action_type,
                        action_text=action_text,
                        status='new',
                        check_date=today,
                    )
                    created_count += 1
                else:
                    # Заявка активна – обновляем её
                    existing_issue.model_1c = model_1c
                    existing_issue.device = device
                    existing_issue.model_db = model_db
                    existing_issue.model_match_type = model_match_type
                    existing_issue.is_online = is_online
                    existing_issue.last_reading_date = last_reading_date
                    existing_issue.last_reading_value = last_reading_value
                    existing_issue.region = default_region
                    existing_issue.action_type = action_type
                    existing_issue.action_text = action_text
                    existing_issue.check_date = today  # обновляем дату проверки
                    existing_issue.save()
                    updated_count += 1
            else:
                # Нет заявки – создаём новую
                CompareResult.objects.create(
                    serial_1c=serial_1c,
                    model_1c=model_1c,
                    device=device,
                    model_db=model_db,
                    model_match_type=model_match_type,
                    is_online=is_online,
                    last_reading_date=last_reading_date,
                    last_reading_value=last_reading_value,
                    region=default_region,
                    action_type=action_type,
                    action_text=action_text,
                    status='new',
                    check_date=today,
                )
                created_count += 1

        messages.success(
            request,
            f'Загрузка завершена.\n'
            f'📌 Район: {default_region.name}\n'
            f'✅ Создано новых заявок: {created_count}\n'
            f'🔄 Обновлено заявок: {updated_count}\n'
            f'⏭️ Пропущено (OK): {skipped_ok}\n'
            f'⏭️ Пропущено (частичное совпадение): {skipped_partial}'
        )
        return redirect('compare_1c_issues')

    # GET – форма загрузки
    regions = Region.objects.all().order_by('name')
    return render(request, 'meters/compare_1c.html', {'regions': regions})

# @login_required
# @user_passes_test(is_allowed_user)
def compare_1c_export(request):
    """Экспорт результатов сверки в Excel."""
    results = request.session.get('compare_results', [])
    if not results:
        messages.error(request, 'Нет данных для экспорта.')
        return redirect('compare_1c')

    wb = Workbook()
    ws = wb.active
    ws.title = "Сверка с 1С"

    headers = [
        '№', 'Заводской номер (1С)', 'Найден в базе', 'Тип ПУ (1С)', 'Тип ПУ (база)',
        'Совпадение модели', 'Последнее показание', 'Дата показания', 'Статус связи',
        'Район', 'Действие'
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='2d3e50', end_color='2d3e50', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    for row_idx, item in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=row_idx-1)
        ws.cell(row=row_idx, column=2, value=item['serial_1c'])
        ws.cell(row=row_idx, column=3, value='Да' if item['found'] else 'Нет')
        ws.cell(row=row_idx, column=4, value=item['model_1c'])
        ws.cell(row=row_idx, column=5, value=item['model_db'])
        # Совпадение модели
        match_display = {
            'exact': 'Полное',
            'partial': 'Частичное',
            'none': 'Не совпадает'
        }.get(item.get('model_match_type'), '—')
        ws.cell(row=row_idx, column=6, value=match_display)
        ws.cell(row=row_idx, column=7, value=item['last_reading_value'])
        ws.cell(row=row_idx, column=8, value=item['last_reading_date'] or '')
        ws.cell(row=row_idx, column=9, value='Онлайн' if item['is_online'] else 'Оффлайн')
        ws.cell(row=row_idx, column=10, value=item['region'])
        ws.cell(row=row_idx, column=11, value=item.get('action_text', ''))

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="compare_1c_{timezone.now().date()}.xlsx"'
    wb.save(response)
    return response

from django.http import StreamingHttpResponse
import csv

# @login_required
# @user_passes_test(is_allowed_user)
def export_devices_csv(request):
    """Экспорт реестра приборов в CSV (потоково, без ограничений)."""
    # Берём те же фильтры, что и в devices
    qs = Device.objects.select_related('model', 'region', 'substation', 'feeder', 'tp').prefetch_related('groups')

    # Применяем фильтры (копируем из devices)
    search = request.GET.get('search', '')
    producer_filter = request.GET.get('producer', '')
    model_filter = request.GET.get('model', '')
    status_filter = request.GET.get('status', '')
    region_filter = request.GET.get('region', '')
    substation_filter = request.GET.get('substation', '')
    feeder_filter = request.GET.get('feeder', '')
    tp_filter = request.GET.get('tp', '')
    group_filter = request.GET.get('group', '')

    if search:
        qs = qs.filter(Q(serial_number__icontains=search) | Q(model__model_name__icontains=search))
    if producer_filter:
        qs = qs.filter(model__device_type_str=producer_filter)
    if model_filter:
        qs = qs.filter(model__id=model_filter)
    if status_filter:
        qs = qs.filter(status=status_filter)
    if region_filter:
        qs = qs.filter(region_id=region_filter)
    if substation_filter:
        qs = qs.filter(substation_id=substation_filter)
    if feeder_filter:
        qs = qs.filter(feeder_id=feeder_filter)
    if tp_filter:
        qs = qs.filter(tp_id=tp_filter)
    if group_filter:
        qs = qs.filter(groups__id=group_filter)

    # Определяем онлайн-статус для каждого устройства
    # Аннотируем has_recent_reading (как в devices)
    seven_days_ago = timezone.now() - timedelta(days=7)
    qs = qs.annotate(
        has_recent_reading=Exists(
            Reading.objects.filter(
                device=OuterRef('pk'),
                timestamp__gte=seven_days_ago
            )
        )
    )

    # Генератор строк CSV
    def generate_csv():
        # Заголовки
        yield ['Заводской номер', 'Тип ПУ', 'Статус', 'Категория', 'Район', 'Подстанция', 'Фидер', 'ТП', 'Группы', 'Ток', 'Фазность', 'Статус связи', 'Дата добавления']

        for device in qs.iterator(chunk_size=1000):
            groups = ', '.join(g.name for g in device.groups.all())
            online_status = 'Онлайн' if device.has_recent_reading else 'Оффлайн'
            yield [
                device.serial_number,
                device.model.model_name if device.model else '',
                device.status,
                device.get_category_display() or '',
                device.region.name if device.region else '',
                device.substation.name if device.substation else '',
                device.feeder.name if device.feeder else '',
                device.tp.name if device.tp else '',
                groups,
                device.nominal_current or '',
                device.model.phases if device.model else '',
                online_status,
                device.created_at.strftime('%Y-%m-%d %H:%M:%S') if device.created_at else '',
            ]

    # Создаём StreamingHttpResponse
    response = StreamingHttpResponse(
        generate_csv(),
        content_type='text/csv; charset=utf-8'
    )
    response['Content-Disposition'] = f'attachment; filename="registry_devices_{timezone.now().date()}.csv"'
    return response

from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, get_object_or_404, redirect
from django.core.paginator import Paginator
from django.contrib import messages
from django.db.models import Q
from .models import CompareResult, Device, Region, Reading
import pandas as pd
import re

# @login_required
# @user_passes_test(is_allowed_user)
def compare_1c_issues(request):
    """Страница со списком заявок на исправление (карточки)."""
    # Фильтры
    status_filter = request.GET.get('status', 'new')  # по умолчанию новые
    region_filter = request.GET.get('region', '')
    action_filter = request.GET.get('action', '')
    upload_region = request.GET.get('upload_region', '')
    verified_filter = request.GET.get('verified', '')  # '' - все, 'yes', 'no'

    # Базовый QuerySet – только записи, требующие действия (исключаем ok и partial)
    qs = CompareResult.objects.exclude(
        action_type__in=['ok']
    ).exclude(
        model_match_type='partial'
    ).select_related('device', 'region', 'fixed_by')

    if status_filter:
        qs = qs.filter(status=status_filter)
    if region_filter:
        qs = qs.filter(region_id=region_filter)
    if action_filter:
        qs = qs.filter(action_type=action_filter)
    if verified_filter == 'yes':
        qs = qs.filter(verified=True)
    elif verified_filter == 'no':
        qs = qs.filter(verified=False)

    # Сортировка: сначала новые, потом по дате создания
    qs = qs.order_by('-updated_at')

    paginator = Paginator(qs, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Подготовка фильтров для выпадающих списков
    status_choices = [
        ('new', 'Новые', status_filter == 'new'),
        ('in_progress', 'в процессе', status_filter == 'in_progress'),
        ('fixed', 'Исправленные', status_filter == 'fixed'),
        ('auto_fixed', 'Автозакрыто', status_filter == 'auto_fixed'),
        ('ignored', 'Пропущенные', status_filter == 'ignored'),
        ('', 'Все', status_filter == ''),
    ]

    action_choices = [
        ('', 'Все', action_filter == ''),
        ('fix_model', 'Исправить Тип ПУ', action_filter == 'fix_model'),
        ('fix_serial', 'Исправить номер', action_filter == 'fix_serial'),
        ('not_found', 'Не найден', action_filter == 'not_found'),
        ('offline', 'Проверить связь', action_filter == 'offline'),
        ('temporary_off', 'Временно отключен', action_filter == 'temporary_off'),
    ]

    verified_choices = [
        ('', 'Все', verified_filter == ''),
        ('yes', 'Проверенные', verified_filter == 'yes'),
        ('no', 'Не проверенные', verified_filter == 'no'),
    ]

    can_verify = request.user.is_staff or request.user.groups.filter(name='operator').exists()

    regions = Region.objects.all().order_by('name')
    region_choices = [(r.id, r.name, str(r.id) == region_filter) for r in regions]
    region_choices.insert(0, ('', 'Все', region_filter == ''))

    context = {
        'page_obj': page_obj,
        'total_count': qs.count(),
        'status_choices': status_choices,
        'action_choices': action_choices,
        'region_choices': region_choices,
        'selected_status': status_filter,
        'selected_action': action_filter,
        'selected_region': region_filter,
        'verified_choices': verified_choices,
        'selected_verified': verified_filter,
        'can_verify': can_verify,
    }
    return render(request, 'meters/compare_1c_issues.html', context)


# @login_required
# @user_passes_test(is_allowed_user)
def compare_1c_issue_update(request, issue_id):
    issue = get_object_or_404(CompareResult, id=issue_id)
    if request.method == 'POST':
        new_status = request.POST.get('status')
        comment = request.POST.get('comment', '').strip()
    
        if new_status in dict(CompareResult.STATUS_CHOICES):
            issue.status = new_status
            
            # Устанавливаем время обработки для всех статусов, кроме 'new'
            if new_status != 'new':
                issue.fixed_at = timezone.now()
                if request.user.is_authenticated:
                    issue.fixed_by = request.user
                else:
                    issue.fixed_by = None
            else:
                # Если случайно выбрали 'new' (хотя мы его скрыли), сбрасываем время
                issue.fixed_at = None
                issue.fixed_by = None

            if comment:
                issue.comment = comment
            issue.save()
            messages.success(request, f'Заявка обновлена: {issue.get_status_display()}')
        else:
            messages.error(request, 'Некорректный статус')
        return redirect('compare_1c_issues')

    status_choices_with_selected = [
        (value, label, issue.status == value) for value, label in CompareResult.STATUS_CHOICES if value not in ('new', 'auto_fixed')
    ]
    context = {
        'issue': issue,
        'status_choices': status_choices_with_selected,
    }
    return render(request, 'meters/compare_1c_issue_update.html', context)


# @login_required
# @user_passes_test(is_allowed_user)
def compare_1c_issues_export(request):
    """Экспорт заявок в CSV."""
    import csv
    from django.http import StreamingHttpResponse

    qs = CompareResult.objects.exclude(
        action_type__in=['ok']
    ).exclude(
        model_match_type='partial'
    ).select_related('device', 'region')

    def generate():
        yield ['ID', 'Заводской номер (1С)', 'Тип ПУ (1С)', 'Тип ПУ (база)', 'Совпадение модели',
               'Район', 'Статус связи', 'Действие', 'Статус заявки', 'Комментарий', 'Дата создания']
        for item in qs.iterator():
            yield [
                item.id,
                item.serial_1c,
                item.model_1c,
                item.model_db,
                item.get_model_match_type_display(),
                item.region.name if item.region else '',
                'Онлайн' if item.is_online else 'Оффлайн',
                item.action_text,
                item.get_status_display(),
                item.comment,
                item.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            ]

    response = StreamingHttpResponse(generate(), content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="compare_issues_{timezone.now().date()}.csv"'
    return response

# @login_required
# @user_passes_test(is_allowed_user)
import json
def compare_1c_stats(request):
    region_filter = request.GET.get('region', '')
    
    qs = CompareResult.objects.exclude(
        Q(action_type='ok') | Q(model_match_type='partial')
    )
    
    if region_filter:
        qs = qs.filter(region_id=region_filter)
    
    # Сначала получаем все статусы
    total_stats = qs.aggregate(
        total=Count('id'),
        new=Count('id', filter=Q(status='new')),
        in_progress=Count('id', filter=Q(status='in_progress')),
        fixed=Count('id', filter=Q(status='fixed')),
        auto_fixed=Count('id', filter=Q(status='auto_fixed')),
        ignored=Count('id', filter=Q(status='ignored')),
    )
    
    # Оффлайн – считаем отдельно: это заявки с action_type='offline' И status != 'auto_fixed' (исключаем уже автозакрытые)
    # Также исключаем fixed и ignored, которые не являются оффлайн
    offline_count = qs.filter(
        action_type='offline',
        status__in=['new', 'in_progress']  # только активные оффлайн
    ).count()
    
    # Закрытые = исправлено + автозакрыто (фиксированные вручную или автоматически)
    closed_total = total_stats['fixed'] + total_stats['auto_fixed']
    
    # Процент закрытых от общего числа
    if total_stats['total'] > 0:
        closed_percent = round((closed_total / total_stats['total']) * 100, 1)
    else:
        closed_percent = 0
    
    # Статистика по районам
    region_stats = qs.values('region__id', 'region__name').annotate(
        total=Count('id'),
        new=Count('id', filter=Q(status='new')),
        in_progress=Count('id', filter=Q(status='in_progress')),
        fixed=Count('id', filter=Q(status='fixed')),
        auto_fixed=Count('id', filter=Q(status='auto_fixed')),
        ignored=Count('id', filter=Q(status='ignored')),
        offline=Count('id', filter=Q(action_type='offline', status__in=['new', 'in_progress'])),
    ).order_by('region__name')
    
    # Данные для диаграмм (если они используются)
    pie_data = {
        'labels': ['Новые', 'в процессе', 'Исправлено', 'Автозакрыто', 'Пропущено'],
        'values': [
            total_stats['new'],
            total_stats['in_progress'],
            total_stats['fixed'],
            total_stats['auto_fixed'],
            total_stats['ignored'],
        ],
        'colors': ['#ffc107', '#0d6efd', '#198754', '#17a2b8', '#6c757d'],
    }
    
    bar_data = {
        'labels': [s['region__name'] or 'Без района' for s in region_stats],
        'total': [s['total'] for s in region_stats],
        'new': [s['new'] for s in region_stats],
        'offline': [s['offline'] for s in region_stats],
    }
    
    regions = Region.objects.all().order_by('name')
    region_choices = [(r.id, r.name, str(r.id) == region_filter) for r in regions]
    region_choices.insert(0, ('', 'Все районы', region_filter == ''))

    context = {
        'total_stats': {
            'total': total_stats['total'],
            'new': total_stats['new'],
            'in_progress': total_stats['in_progress'],
            'fixed': total_stats['fixed'],
            'auto_fixed': total_stats['auto_fixed'],
            'ignored': total_stats['ignored'],
            'offline': offline_count,
            'closed': closed_total,
            'closed_percent': closed_percent,
        },
        'region_stats': region_stats,
        'pie_data_json': json.dumps(pie_data),
        'bar_data_json': json.dumps(bar_data),
        'region_choices': region_choices,
        'selected_region': region_filter,
        'has_results': region_stats.exists(),
    }
    return render(request, 'meters/compare_1c_stats.html', context)

@login_required
@user_passes_test(is_allowed_user)
def export_devices_xlsx(request):
    """Экспорт реестра приборов в Excel (только до 5000 записей, иначе предлагаем CSV)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from django.db.models import Exists, OuterRef

    # Базовый queryset с фильтрами (как в devices)
    qs = Device.objects.select_related('model', 'region', 'substation', 'feeder', 'tp').prefetch_related('groups')
    seven_days_ago = timezone.now() - timedelta(days=7)
    qs = qs.annotate(
        has_recent_reading=Exists(
            Reading.objects.filter(
                device=OuterRef('pk'),
                timestamp__gte=seven_days_ago
            )
        )
    )

    # Фильтры (копируем из devices)
    search = request.GET.get('search', '')
    producer_filter = request.GET.get('producer', '')
    model_filter = request.GET.get('model', '')
    status_filter = request.GET.get('status', '')
    region_filter = request.GET.get('region', '')
    substation_filter = request.GET.get('substation', '')
    feeder_filter = request.GET.get('feeder', '')
    tp_filter = request.GET.get('tp', '')
    group_filter = request.GET.get('group', '')

    if search:
        qs = qs.filter(Q(serial_number__icontains=search) | Q(model__model_name__icontains=search))
    if producer_filter:
        qs = qs.filter(model__device_type_str=producer_filter)
    if model_filter:
        qs = qs.filter(model__id=model_filter)
    if status_filter:
        qs = qs.filter(status=status_filter)
    if region_filter:
        qs = qs.filter(region_id=region_filter)
    if substation_filter:
        qs = qs.filter(substation_id=substation_filter)
    if feeder_filter:
        qs = qs.filter(feeder_id=feeder_filter)
    if tp_filter:
        qs = qs.filter(tp_id=tp_filter)
    if group_filter:
        qs = qs.filter(groups__id=group_filter)

    count = qs.count()
    MAX_ROWS = 10000
    if count > MAX_ROWS:
        messages.warning(request, f'Слишком много записей ({count}). Используйте CSV-экспорт.')
        return redirect('devices')

    # Создаём книгу
    wb = Workbook()
    ws = wb.active
    ws.title = "Реестр приборов"

    headers = [
        'Заводской номер', 'Тип ПУ', 'Статус', 'Категория', 'Район', 'Подстанция',
        'Фидер', 'ТП', 'Группы', 'Ток', 'Фазность', 'Статус связи', 'Дата добавления'
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='2d3e50', end_color='2d3e50', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    for row_idx, device in enumerate(qs, 2):
        groups = ', '.join(g.name for g in device.groups.all())
        online_status = 'Онлайн' if device.has_recent_reading else 'Оффлайн'
        ws.cell(row=row_idx, column=1, value=device.serial_number)
        ws.cell(row=row_idx, column=2, value=device.model.model_name if device.model else '')
        ws.cell(row=row_idx, column=3, value=device.status)
        ws.cell(row=row_idx, column=4, value=device.get_category_display())
        ws.cell(row=row_idx, column=5, value=device.region.name if device.region else '')
        ws.cell(row=row_idx, column=6, value=device.substation.name if device.substation else '')
        ws.cell(row=row_idx, column=7, value=device.feeder.name if device.feeder else '')
        ws.cell(row=row_idx, column=8, value=device.tp.name if device.tp else '')
        ws.cell(row=row_idx, column=9, value=groups)
        ws.cell(row=row_idx, column=10, value=device.nominal_current or '')
        ws.cell(row=row_idx, column=11, value=device.model.phases if device.model else '')
        ws.cell(row=row_idx, column=12, value=online_status)
        ws.cell(row=row_idx, column=13, value=device.created_at.strftime('%Y-%m-%d %H:%M:%S') if device.created_at else '')

    # Автоширина
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # Сохраняем
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="registry_devices_{timezone.now().date()}.xlsx"'
    return response

from django.http import JsonResponse
from django.core.management import call_command
from django.contrib.auth.decorators import login_required

@login_required
def check_offline_issues(request):
    if not request.user.is_staff:
        return JsonResponse({'error': 'Требуются права администратора'}, status=403)
    try:
        call_command('auto_close_issues')
        return JsonResponse({'message': 'Автозакрытие выполнено успешно'}, status=200)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse

def compare_1c_issues_export_excel(request):
    """Экспорт заявок в Excel."""
    qs = CompareResult.objects.exclude(
        action_type__in=['ok']
    ).exclude(
        model_match_type='partial'
    ).select_related('device', 'region')

    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки на исправление"

    headers = [
        'ID', 'Заводской номер (1С)', 'Тип ПУ (1С)', 'Тип ПУ (база)',
        'Совпадение модели', 'Район', 'Статус связи', 'Действие',
        'Статус заявки', 'Комментарий', 'Дата создания', 'Дата обработки'
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='2d3e50', end_color='2d3e50', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    for row_idx, item in enumerate(qs, 2):
        ws.cell(row=row_idx, column=1, value=item.id)
        ws.cell(row=row_idx, column=2, value=item.serial_1c)
        ws.cell(row=row_idx, column=3, value=item.model_1c)
        ws.cell(row=row_idx, column=4, value=item.model_db)
        ws.cell(row=row_idx, column=5, value=item.get_model_match_type_display())
        ws.cell(row=row_idx, column=6, value=item.region.name if item.region else '')
        ws.cell(row=row_idx, column=7, value='Онлайн' if item.is_online else 'Оффлайн')
        ws.cell(row=row_idx, column=8, value=item.action_text)
        ws.cell(row=row_idx, column=9, value=item.get_status_display())
        ws.cell(row=row_idx, column=10, value=item.comment)
        ws.cell(row=row_idx, column=11, value=item.created_at.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row_idx, column=12, value=item.fixed_at.strftime('%Y-%m-%d %H:%M:%S') if item.fixed_at else '')

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="issues_{timezone.now().date()}.xlsx"'
    wb.save(response)
    return response

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required

@login_required
def compare_1c_issue_verify(request, issue_id):
    """Переключает отметку 'Проверено' для заявки (только для staff/operator)."""
    if not request.user.is_staff and not request.user.groups.filter(name='operator').exists():
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    issue = get_object_or_404(CompareResult, id=issue_id)
    if request.method == 'POST':
        issue.verified = not issue.verified
        issue.save()
        return JsonResponse({'verified': issue.verified, 'id': issue.id})
    return JsonResponse({'error': 'Method not allowed'}, status=405)

from django.http import JsonResponse
from django.views.decorators.http import require_POST
from .models import Device, ReadingFetchTask
import threading

@require_POST
@login_required
def fetch_device_readings(request):
    serial_number = request.POST.get('serial_number')
    if not serial_number:
        return JsonResponse({'error': 'Не указан заводской номер'}, status=400)

    try:
        device = Device.objects.get(serial_number=serial_number)
    except Device.DoesNotExist:
        return JsonResponse({'error': 'Устройство не найдено'}, status=404)

    # Проверяем, есть ли уже запущенная задача для этого устройства
    existing_task = ReadingFetchTask.objects.filter(
        device=device,
        status__in=['pending', 'running']
    ).first()
    if existing_task:
        return JsonResponse({
            'task_id': existing_task.id,
            'status': existing_task.status,
            'message': 'Задача уже выполняется'
        })

    # Создаём задачу
    task = ReadingFetchTask.objects.create(
        device=device,
        serial_number=serial_number,
        status='pending'
    )

    # Запускаем в фоновом потоке
    def run_task():
        from django.core.management import call_command
        call_command('fetch_readings_task', str(task.id))

    thread = threading.Thread(target=run_task)
    thread.daemon = True
    thread.start()

    return JsonResponse({
        'task_id': task.id,
        'status': 'pending',
        'message': 'Задача запущена'
    })

@login_required
def fetch_readings_status(request):
    task_id = request.GET.get('task_id')
    if not task_id:
        return JsonResponse({'error': 'Не указан task_id'}, status=400)

    try:
        task = ReadingFetchTask.objects.get(id=task_id)
    except ReadingFetchTask.DoesNotExist:
        return JsonResponse({'error': 'Задача не найдена'}, status=404)

    return JsonResponse({
        'status': task.status,
        'records_loaded': task.records_loaded,
        'error_message': task.error_message,
        'completed_at': task.completed_at.isoformat() if task.completed_at else None,
    })

def category_stats_table(request):
    """Возвращает HTML-таблицу со статистикой по категориям для выбранного района."""
    region_id = request.GET.get('region_id')
    period = request.GET.get('period', 'today')
    cache_key = f'category_stats_{period}_{region_id}'
    html = cache.get(cache_key)
    if html is None:
        today = timezone.now().date()
        if period == 'today':
            start_date = today
        elif period == '3d':
            start_date = today - timedelta(days=3)
        elif period == '7d':
            start_date = today - timedelta(days=7)
        elif period == '30d':
            start_date = today - timedelta(days=30)
        else:
            start_date = today

        start_dt = datetime.combine(start_date, datetime.min.time())
        end_dt = datetime.combine(today, datetime.max.time())

        devices = Device.objects.filter(status='active')
        if region_id:
            devices = devices.filter(region_id=region_id)

        # Аннотируем has_reading
        from django.db.models import Exists, OuterRef, Count, Case, When, IntegerField, Value
        readings_subquery = Reading.objects.filter(
            device=OuterRef('pk'),
            timestamp__gte=start_dt,
            timestamp__lte=end_dt
        ).values('device_id').annotate(cnt=Count('id')).values('cnt')
        devices = devices.annotate(
            has_reading=Exists(readings_subquery)
        )

        # Группируем по категориям
        category_stats = devices.values('category').annotate(
            total=Count('id'),
            active=Count(Case(When(has_reading=True, then=1), output_field=IntegerField()))
        ).order_by('category')

        # Преобразуем для шаблона
        stats = []
        for stat in category_stats:
            cat = stat['category']
            total = stat['total']
            active = stat['active']
            percent = round((active / total * 100), 1) if total else 0
            stats.append({
                'category': cat,
                'category_display': dict(Device.CATEGORY_CHOICES).get(cat, cat or 'Не указана'),
                'total': total,
                'active': active,
                'percent': percent,
            })

        html = render_to_string('meters/_category_table.html', {'category_stats': stats})
        cache.set(cache_key, html, 600)
    return HttpResponse(html)

@login_required
@user_passes_test(is_allowed_user)
def bulk_update_category_page(request):
    if request.method == 'POST':
        serials_text = request.POST.get('serials', '').strip()
        new_category = request.POST.get('category', '')
        if not serials_text:
            messages.error(request, 'Список серийных номеров пуст.')
            return redirect('bulk_update_category_page')
        if not new_category:
            messages.error(request, 'Выберите категорию.')
            return redirect('bulk_update_category_page')
        serials = [s.strip() for s in serials_text.split('\n') if s.strip()]
        updated = 0
        not_found = []
        for sn in serials:
            device = Device.objects.filter(serial_number=sn).first()
            if device:
                device.category = new_category
                device.save()
                updated += 1
            else:
                not_found.append(sn)
        if updated:
            messages.success(request, f'Обновлено устройств: {updated}')
        if not_found:
            messages.warning(request, f'Не найдены: {", ".join(not_found[:10])}')
        return redirect('bulk_update_category_page')
    # GET – показываем форму
    categories = Device.CATEGORY_CHOICES
    return render(request, 'meters/bulk_update_category.html', {'categories': categories})

# Главная страница иерархии
@login_required
@user_passes_test(is_allowed_user)
def hierarchy_management(request):
    """Страница управления иерархией в виде дерева."""
    tree = get_hierarchy_tree()
    context = {
        'tree': tree,
    }
    return render(request, 'meters/hierarchy.html', context)

def get_hierarchy_tree():
    tree = []
    regions = Region.objects.all().order_by('name')
    for region in regions:
        region_node = {
            'id': region.id,
            'name': region.name,
            'type': 'region',
            'children': []
        }
        substations = region.substations.all().order_by('name')
        for substation in substations:
            substation_node = {
                'id': substation.id,
                'name': substation.name,
                'type': 'substation',
                'children': []
            }
            # Вводные фидеры (parent_feeder is null, feeder_type='input')
            input_feeders = substation.feeders.filter(parent_feeder__isnull=True, feeder_type='input').order_by('name')
            for feeder in input_feeders:
                feeder_node = {
                    'id': feeder.id,
                    'name': feeder.name,
                    'type': 'feeder',
                    'feeder_type': feeder.get_feeder_type_display(),
                    'head_meter': feeder.head_meter.serial_number if feeder.head_meter else None,
                    'children': []
                }
                # Отходящие фидеры (parent_feeder = this feeder)
                output_feeders = feeder.children.all().order_by('name')
                for out_feeder in output_feeders:
                    out_feeder_node = {
                        'id': out_feeder.id,
                        'name': out_feeder.name,
                        'type': 'feeder',
                        'feeder_type': out_feeder.get_feeder_type_display(),
                        'head_meter': out_feeder.head_meter.serial_number if out_feeder.head_meter else None,
                        'children': []
                    }
                    # ТП привязаны к отходящему фидеру (или можно к любому, но обычно к отходящему)
                    tps = out_feeder.tps.all().order_by('name')
                    for tp in tps:
                        tp_node = {
                            'id': tp.id,
                            'name': tp.name,
                            'type': 'tp',
                            'head_meter': tp.head_meter.serial_number if tp.head_meter else None,
                            'children': []
                        }
                        out_feeder_node['children'].append(tp_node)
                    feeder_node['children'].append(out_feeder_node)
                substation_node['children'].append(feeder_node)
            region_node['children'].append(substation_node)
        tree.append(region_node)
    return tree

@login_required
@user_passes_test(is_allowed_user)
def hierarchy_node_details(request):
    node_id = request.GET.get('id')
    node_type = request.GET.get('type')
    
    context = {}
    if node_type == 'region':
        obj = get_object_or_404(Region, id=node_id)
        context['obj'] = obj
        context['type'] = 'region'
        context['label'] = 'Район'
        context['parent_label'] = None
        context['parent_choices'] = []
        context['feeder_type_choices'] = []
        context['show_head_meter'] = False  # у района нет головного
    elif node_type == 'substation':
        obj = get_object_or_404(Substation, id=node_id)
        context['obj'] = obj
        context['type'] = 'substation'
        context['label'] = 'Подстанция'
        context['parent_label'] = 'Район'
        parents = Region.objects.all().order_by('name')
        context['parent_choices'] = [(p.id, p.name, p.id == obj.region_id) for p in parents]
        context['feeder_type_choices'] = []
        context['show_head_meter'] = False  # у подстанции нет головного
    elif node_type == 'feeder':
        obj = get_object_or_404(Feeder, id=node_id)
        context['obj'] = obj
        context['type'] = 'feeder'
        context['label'] = 'Фидер'
        # Если фидер вводной, то родитель — подстанция
        if obj.feeder_type == 'input':
            context['parent_label'] = 'Подстанция'
            parents = Substation.objects.all().order_by('name')
            context['parent_choices'] = [(p.id, p.name, p.id == obj.substation_id) for p in parents]
            context['parent_field'] = 'substation_id'  # для формы
            context['show_parent_feeder'] = False
        else:  # отходящий
            context['parent_label'] = 'Вводной фидер'
            # Список вводных фидеров той же подстанции
            input_feeders = Feeder.objects.filter(substation=obj.substation, feeder_type='input', parent_feeder__isnull=True).order_by('name')
            context['parent_choices'] = [(f.id, f.name, f.id == obj.parent_feeder_id) for f in input_feeders]
            context['parent_field'] = 'parent_feeder_id'
            context['show_parent_feeder'] = True
        context['feeder_type_choices'] = [
            ('input', 'Вводной', obj.feeder_type == 'input'),
            ('output', 'Отходящий', obj.feeder_type == 'output'),
        ]
        context['show_head_meter'] = True  # у фидера есть головной
    elif node_type == 'tp':
        obj = get_object_or_404(TransformerSubstation, id=node_id)
        context['obj'] = obj
        context['type'] = 'tp'
        context['label'] = 'ТП'
        context['parent_label'] = 'Фидер'
        parents = Feeder.objects.all().order_by('name')
        context['parent_choices'] = [(p.id, p.name, p.id == obj.feeder_id) for p in parents]
        context['feeder_type_choices'] = []
        context['show_head_meter'] = True  # у ТП может быть головной
    else:
        return HttpResponse('<p class="text-muted">Неизвестный тип</p>')
    
    context['devices'] = Device.objects.filter(status='active').order_by('serial_number')
    return render(request, 'meters/_node_details.html', context)


# управления иерархией: районы → подстанции → фидеры → ТП с возможностью назначения головных счётчиков и массовой привязки устройств к ТП
# AJAX-таблица для вкладок  
# удалить потом
@login_required
@user_passes_test(is_allowed_user)
def hierarchy_table(request):
    level = request.GET.get('level', 'region')
    devices = Device.objects.filter(status='active').order_by('serial_number')

    if level == 'region':
        objects = Region.objects.all().order_by('name')
        columns = ['Название']
        fields = ['name']
        parent_field = None
    elif level == 'substation':
        objects = Substation.objects.select_related('region').all().order_by('name')
        columns = ['Название', 'Район']
        fields = ['name', 'region']
        parent_field = 'region'
        # head_meter убрали
    elif level == 'feeder':
        objects = Feeder.objects.select_related('substation').all().order_by('name')
        columns = ['Название', 'Подстанция', 'Тип', 'Головной счётчик']
        fields = ['name', 'substation', 'feeder_type', 'head_meter']
        parent_field = 'substation'
    elif level == 'tp':
        objects = TransformerSubstation.objects.select_related('feeder').all().order_by('name')
        columns = ['Название', 'Фидер', 'Головной счётчик']
        fields = ['name', 'feeder', 'head_meter']
        parent_field = 'feeder'
    else:
        objects = []
        columns = []
        fields = []
        parent_field = None

    context = {
        'level': level,
        'objects': objects,
        'columns': columns,
        'fields': fields,
        'parent_field': parent_field,
        'devices': devices,
    }
    return render(request, 'meters/_hierarchy_table.html', context)

# Добавление/редактирование объекта (модальное окно)
@login_required
@user_passes_test(is_allowed_user)
def hierarchy_add(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    level = request.POST.get('level')
    name = request.POST.get('name', '').strip()
    head_meter_id = request.POST.get('head_meter_id')
    parent_id = request.POST.get('parent_id')
    feeder_type = request.POST.get('feeder_type')  # для фидера

    if not name:
        return JsonResponse({'error': 'Название обязательно'}, status=400)

    try:
        if level == 'region':
            obj = Region.objects.create(name=name)
        elif level == 'substation':
            if not parent_id:
                return JsonResponse({'error': 'Выберите район'}, status=400)
            obj = Substation.objects.create(
                name=name,
                region_id=parent_id,
            )
        elif level == 'feeder':
            if not parent_id:
                return JsonResponse({'error': 'Выберите подстанцию'}, status=400)
            if not feeder_type:
                return JsonResponse({'error': 'Выберите тип фидера'}, status=400)
            obj = Feeder.objects.create(
                name=name,
                substation_id=parent_id,
                feeder_type=feeder_type,
                head_meter_id=head_meter_id or None
            )
        elif level == 'tp':
            if not parent_id:
                return JsonResponse({'error': 'Выберите фидер'}, status=400)
            obj = TransformerSubstation.objects.create(
                name=name,
                feeder_id=parent_id,
                head_meter_id=head_meter_id or None
            )
        else:
            return JsonResponse({'error': 'Неверный уровень'}, status=400)
        return JsonResponse({'success': True, 'id': obj.id, 'name': obj.name})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@user_passes_test(is_allowed_user)
def hierarchy_edit(request, pk):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    level = request.POST.get('level')
    name = request.POST.get('name', '').strip()
    head_meter_id = request.POST.get('head_meter_id')
    parent_id = request.POST.get('parent_id')
    feeder_type = request.POST.get('feeder_type')

    if not name:
        return JsonResponse({'error': 'Название обязательно'}, status=400)

    try:
        if level == 'region':
            obj = get_object_or_404(Region, id=pk)
            obj.name = name
            obj.save()

        elif level == 'substation':
            if not parent_id:
                return JsonResponse({'error': 'Выберите район'}, status=400)
            obj = get_object_or_404(Substation, id=pk)
            obj.name = name
            obj.region_id = parent_id
            obj.save()

        elif level == 'feeder':
            if not parent_id:
                return JsonResponse({'error': 'Выберите подстанцию'}, status=400)
            if not feeder_type:
                return JsonResponse({'error': 'Выберите тип фидера'}, status=400)
            obj = get_object_or_404(Feeder, id=pk)
            obj.name = name
            obj.substation_id = parent_id
            obj.feeder_type = feeder_type
            obj.head_meter_id = head_meter_id or None
            obj.save()

        elif level == 'tp':
            if not parent_id:
                return JsonResponse({'error': 'Выберите фидер'}, status=400)
            obj = get_object_or_404(TransformerSubstation, id=pk)
            obj.name = name
            obj.feeder_id = parent_id
            obj.head_meter_id = head_meter_id or None
            obj.save()

        else:
            return JsonResponse({'error': 'Неверный уровень'}, status=400)

        return JsonResponse({'success': True})

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# Удаление объекта
@login_required
@user_passes_test(is_allowed_user)
def hierarchy_delete(request, pk):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    level = request.POST.get('level')
    try:
        if level == 'region':
            obj = get_object_or_404(Region, id=pk)
        elif level == 'substation':
            obj = get_object_or_404(Substation, id=pk)
        elif level == 'feeder':
            obj = get_object_or_404(Feeder, id=pk)
        elif level == 'tp':
            obj = get_object_or_404(TransformerSubstation, id=pk)
        else:
            return JsonResponse({'error': 'Неверный уровень'}, status=400)
        obj.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# Массовая привязка устройств к ТП
@login_required
@user_passes_test(is_allowed_user)
def bulk_assign_tp(request):
    if request.method == 'POST':
        serials_text = request.POST.get('serials', '').strip()
        tp_id = request.POST.get('tp_id')
        if not serials_text:
            messages.error(request, 'Список серийных номеров пуст.')
            return redirect('bulk_assign_tp')
        if not tp_id:
            messages.error(request, 'Выберите ТП.')
            return redirect('bulk_assign_tp')
        serials = [s.strip() for s in serials_text.split('\n') if s.strip()]
        updated = 0
        not_found = []
        for sn in serials:
            device = Device.objects.filter(serial_number=sn).first()
            if device:
                device.tp_id = tp_id
                device.save()
                updated += 1
            else:
                not_found.append(sn)
        if updated:
            messages.success(request, f'Обновлено устройств: {updated}')
        if not_found:
            messages.warning(request, f'Не найдены: {", ".join(not_found[:10])}')
        return redirect('bulk_assign_tp')

    # GET – показываем форму
    tps = TransformerSubstation.objects.select_related('feeder__substation__region').all().order_by('name')
    context = {'tps': tps}
    return render(request, 'meters/bulk_assign_tp.html', context)

@login_required
@user_passes_test(is_allowed_user)
def hierarchy_parents(request):
    """Возвращает список родительских объектов для указанного уровня в формате JSON."""
    level = request.GET.get('level')
    if level == 'region':
        items = Region.objects.all().order_by('name')
        data = [{'id': r.id, 'name': r.name} for r in items]
    elif level == 'substation':
        items = Substation.objects.all().order_by('name')
        data = [{'id': s.id, 'name': s.name} for s in items]
    elif level == 'feeder':
        items = Feeder.objects.all().order_by('name')
        data = [{'id': f.id, 'name': f.name} for f in items]
    else:
        return JsonResponse([], safe=False)
    return JsonResponse(data, safe=False)

from django.http import JsonResponse
from django.views.decorators.http import require_GET

@require_GET
def search_devices(request):
    """API для поиска устройств по заводскому номеру (автодополнение)."""
    term = request.GET.get('term', '').strip()
    if not term or len(term) < 2:
        return JsonResponse([], safe=False)
    devices = Device.objects.filter(
        serial_number__icontains=term,
        status='active'
    ).order_by('serial_number')[:20]
    data = [{'id': d.id, 'text': d.serial_number} for d in devices]
    return JsonResponse(data, safe=False)

# ========== БАЛАНС ==========

from .balance_calculator import (
    calculate_balance_for_substation,
    get_or_create_period,
    get_reading_for_date,
    get_coefficient_for_date,
)
from .models import (
    Device, Reading, SyncStatus, MeterModel, Region, 
    Substation, Feeder, TransformerSubstation, DeviceGroup,
    CompareResult, ReadingFetchTask,
    # Новые модели для баланса
    MeterConnection, CoefficientHistory, CalculationPeriod, BalanceResult
)
from decimal import Decimal

@login_required
@user_passes_test(is_allowed_user)
def balance_readings(request):
    substation_id = request.GET.get('substation')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    # Если даты не заданы, используем 1-е число месяца и сегодня
    if not start_date_str:
        start_date = timezone.now().date().replace(day=1)
    else:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except:
            start_date = timezone.now().date().replace(day=1)

    if not end_date_str:
        end_date = timezone.now().date()
    else:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except:
            end_date = timezone.now().date()

    substations = Substation.objects.filter(is_active=True).order_by('name')
    selected_substation = None
    data = []
    substations_with_selected = []

    if substation_id:
        selected_substation = get_object_or_404(Substation, id=substation_id)
        feeders = selected_substation.feeders.all().order_by('feeder_type', 'name')
        for feeder in feeders:
            if not feeder.head_meter:
                continue
            device = feeder.head_meter
            # Для каждого направления
            for direction in ['aplus', 'aminus']:
                dir_display = 'A+' if direction == 'aplus' else 'A-'
                # Получаем показания на начальную и конечную даты
                reading_start = get_reading_for_date(device, start_date, direction)
                reading_end = get_reading_for_date(device, end_date, direction)
                if reading_start is None or reading_end is None:
                    continue
                # Разность
                diff = reading_end - reading_start
                # Коэффициент на конечную дату
                coeff = get_coefficient_for_date(device, end_date)
                # Объём
                energy = diff * coeff
                data.append({
                    'feeder': feeder,
                    'device': device,
                    'direction': dir_display,
                    'reading_start': reading_start,
                    'reading_end': reading_end,
                    'diff': diff,
                    'coefficient': coeff,
                    'energy': energy,
                })

    for sub in substations:
        selected = (selected_substation and selected_substation.id == sub.id)
        substations_with_selected.append((sub, selected))

    # Вычисляем итоги
    total_input = sum(item['energy'] for item in data if item['feeder'].feeder_type == 'input')
    total_output = sum(item['energy'] for item in data if item['feeder'].feeder_type == 'output')
    imbalance = total_input - total_output
    imbalance_percent = (imbalance / total_input * 100) if total_input else Decimal(0)

    context = {
        'substations_with_selected': substations_with_selected,
        'selected_substation': selected_substation,
        'start_date': start_date,
        'end_date': end_date,
        'data': data,
        'total_input': total_input,
        'total_output': total_output,
        'imbalance': imbalance,
        'imbalance_percent': imbalance_percent,
    }
    return render(request, 'meters/balance_readings.html', context)

@login_required
@user_passes_test(is_allowed_user)
def balance_summary(request):
    region_id = request.GET.get('region')
    year = int(request.GET.get('year', timezone.now().year))
    month = int(request.GET.get('month', timezone.now().month))

    substations = Substation.objects.filter(is_active=True)
    if region_id:
        substations = substations.filter(region_id=region_id)

    results = []
    total_decades = {
        'decade1': {'input': 0, 'output': 0, 'imbalance': 0},
        'decade2': {'input': 0, 'output': 0, 'imbalance': 0},
        'decade3': {'input': 0, 'output': 0, 'imbalance': 0},
        'month': {'input': 0, 'output': 0, 'imbalance': 0, 'percent': 0},
    }

    for sub in substations:
        balance = calculate_balance_for_substation(sub, year, month)
        results.append({
            'substation': sub,
            'decades': balance,
        })
        for key in ['decade1', 'decade2', 'decade3', 'month']:
            total_decades[key]['input'] += balance[key]['input']
            total_decades[key]['output'] += balance[key]['output']
            total_decades[key]['imbalance'] += balance[key]['imbalance']
        if total_decades['month']['input'] > 0:
            total_decades['month']['percent'] = (total_decades['month']['imbalance'] / total_decades['month']['input']) * 100

    regions = Region.objects.all().order_by('name')
    regions_with_selected = [(r, str(r.id) == region_id) for r in regions]

    context = {
        'regions_with_selected': regions_with_selected,
        'results': results,
        'year': year,
        'month': month,
        'total_decades': total_decades,
    }
    return render(request, 'meters/balance_summary.html', context)


@login_required
@user_passes_test(is_allowed_user)
def balance_report_all(request):
    year = int(request.GET.get('year', timezone.now().year))
    month = int(request.GET.get('month', timezone.now().month))

    substations = Substation.objects.filter(is_active=True)
    data = []
    total_input = 0
    total_output = 0
    total_imbalance = 0

    for sub in substations:
        period = get_or_create_period(sub, year, month, 'month')
        try:
            balance = BalanceResult.objects.get(substation=sub, period=period)
        except BalanceResult.DoesNotExist:
            calculate_balance_for_substation(sub, year, month)  # расчёт
            balance = BalanceResult.objects.get(substation=sub, period=period)

        data.append({
            'substation': sub,
            'input': balance.input_energy,
            'output': balance.output_energy,
            'imbalance': balance.imbalance_kwh,
            'percent': balance.imbalance_percent,
        })
        total_input += balance.input_energy
        total_output += balance.output_energy
        total_imbalance += balance.imbalance_kwh

    total_percent = (total_imbalance / total_input * 100) if total_input else 0

    chart_labels = [d['substation'].name for d in data]
    chart_imbalance = [float(d['imbalance']) for d in data]
    chart_input = [float(d['input']) for d in data]

    context = {
        'data': data,
        'year': year,
        'month': month,
        'chart_labels': json.dumps(chart_labels),
        'chart_imbalance': json.dumps(chart_imbalance),
        'chart_input': json.dumps(chart_input),
        'total_input': total_input,
        'total_output': total_output,
        'total_imbalance': total_imbalance,
        'total_percent': total_percent,
    }
    return render(request, 'meters/balance_report_all.html', context)


# Вспомогательная функция для получения показания на дату (используется в balance_readings)
# def get_reading_for_date(device, date_obj):
#     from .models import Reading
#     try:
#         reading = Reading.objects.filter(
#             device=device,
#             timestamp__date=date_obj
#         ).order_by('-timestamp').first()
#         return reading.reading_value if reading else None
#     except Reading.DoesNotExist:
#         return None

from decimal import Decimal
from .models import CoefficientHistory

@login_required
@user_passes_test(is_allowed_user)
def device_coefficients(request, device_id):
    device = get_object_or_404(Device, id=device_id)
    
    if request.method == 'POST':
        tt = request.POST.get('tt_ratio')
        tn = request.POST.get('tn_ratio')
        start_date = request.POST.get('start_date')
        
        if not tt or not tn:
            messages.error(request, 'Введите оба коэффициента (ТТ и ТН)')
            return redirect('device_coefficients', device_id=device.id)
        
        try:
            tt = Decimal(tt)
            tn = Decimal(tn)
        except:
            messages.error(request, 'Введите корректные числовые значения')
            return redirect('device_coefficients', device_id=device.id)
        
        if not start_date:
            start_date = timezone.now().date()
        else:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            except:
                start_date = timezone.now().date()
        
        # Закрываем предыдущую активную запись (если есть)
        active_history = CoefficientHistory.objects.filter(
            device=device,
            end_date__isnull=True
        ).first()
        if active_history:
            active_history.end_date = start_date
            active_history.save()
        
        # Создаём новую запись
        CoefficientHistory.objects.create(
            device=device,
            tt_ratio=tt,
            tn_ratio=tn,
            start_date=start_date,
            end_date=None
        )
        
        # Обновляем текущие значения в Device
        device.tt_ratio = tt
        device.tn_ratio = tn
        device.save()
        
        messages.success(request, f'Коэффициенты обновлены (ТТ={tt}, ТН={tn}, с {start_date})')
        return redirect('device_coefficients', device_id=device.id)
    
    # GET – показываем форму и историю
    history = CoefficientHistory.objects.filter(device=device).order_by('-start_date')
    
    # ВЫЧИСЛЯЕМ КОЭФФИЦИЕНТ ДЛЯ КАЖДОЙ ЗАПИСИ
    history_with_calc = []
    for item in history:
        # Добавляем вычисленное поле calc_coeff
        item.calc_coeff = item.tt_ratio * item.tn_ratio
        history_with_calc.append(item)
    
    context = {
        'device': device,
        'history': history_with_calc,  # передаём уже с вычисленным коэффициентом
    }
    return render(request, 'meters/device_coefficients.html', context)
    
# # ===== СПРАВОЧНИКИ (единая страница) =====
# @login_required
# @user_passes_test(is_allowed_user)
# def directories(request):
#     from django.shortcuts import get_object_or_404

#     # Определяем уровень из GET или POST
#     level = request.GET.get('level', 'region')  # region, substation, feeder, tp

#     # Списки для выпадающих полей
#     regions = Region.objects.all().order_by('name')
#     substations_all = Substation.objects.all().order_by('name')
#     feeders_all = Feeder.objects.all().order_by('name')
#     tps_all = TransformerSubstation.objects.all().order_by('name')
#     devices = Device.objects.filter(status='active').order_by('serial_number')

#     # Обработка POST (добавление/редактирование/удаление)
#     if request.method == 'POST':
#         action = request.POST.get('action')
#         level = request.POST.get('level', 'region')

#         if action == 'add' or action == 'edit':
#             # Общие поля
#             obj_id = request.POST.get('id')
#             name = request.POST.get('name', '').strip()
#             head_meter_id = request.POST.get('head_meter_id')

#             if not name:
#                 messages.error(request, 'Название обязательно.')
#                 return redirect(f'{request.path}?level={level}')

#             # В зависимости от уровня собираем родительский объект
#             if level == 'region':
#                 if obj_id:
#                     obj = get_object_or_404(Region, id=obj_id)
#                     obj.name = name
#                     obj.head_meter_id = head_meter_id if head_meter_id else None
#                     obj.save()
#                     messages.success(request, f'Район "{obj.name}" обновлён.')
#                 else:
#                     obj = Region.objects.create(name=name, head_meter_id=head_meter_id or None)
#                     messages.success(request, f'Район "{obj.name}" добавлен.')

#             elif level == 'substation':
#                 region_id = request.POST.get('region_id')
#                 if not region_id:
#                     messages.error(request, 'Выберите район.')
#                     return redirect(f'{request.path}?level={level}')
#                 if obj_id:
#                     obj = get_object_or_404(Substation, id=obj_id)
#                     obj.name = name
#                     obj.region_id = region_id
#                     obj.head_meter_id = head_meter_id or None
#                     obj.save()
#                     messages.success(request, f'Подстанция "{obj.name}" обновлена.')
#                 else:
#                     obj = Substation.objects.create(
#                         name=name,
#                         region_id=region_id,
#                         head_meter_id=head_meter_id or None
#                     )
#                     messages.success(request, f'Подстанция "{obj.name}" добавлена.')

#             elif level == 'feeder':
#                 substation_id = request.POST.get('substation_id')
#                 if not substation_id:
#                     messages.error(request, 'Выберите подстанцию.')
#                     return redirect(f'{request.path}?level={level}')
#                 if obj_id:
#                     obj = get_object_or_404(Feeder, id=obj_id)
#                     obj.name = name
#                     obj.substation_id = substation_id
#                     obj.head_meter_id = head_meter_id or None
#                     obj.save()
#                     messages.success(request, f'Фидер "{obj.name}" обновлён.')
#                 else:
#                     obj = Feeder.objects.create(
#                         name=name,
#                         substation_id=substation_id,
#                         head_meter_id=head_meter_id or None
#                     )
#                     messages.success(request, f'Фидер "{obj.name}" добавлен.')

#             elif level == 'tp':
#                 feeder_id = request.POST.get('feeder_id')
#                 if not feeder_id:
#                     messages.error(request, 'Выберите фидер.')
#                     return redirect(f'{request.path}?level={level}')
#                 if obj_id:
#                     obj = get_object_or_404(TransformerSubstation, id=obj_id)
#                     obj.name = name
#                     obj.feeder_id = feeder_id
#                     obj.head_meter_id = head_meter_id or None
#                     obj.save()
#                     messages.success(request, f'ТП "{obj.name}" обновлена.')
#                 else:
#                     obj = TransformerSubstation.objects.create(
#                         name=name,
#                         feeder_id=feeder_id,
#                         head_meter_id=head_meter_id or None
#                     )
#                     messages.success(request, f'ТП "{obj.name}" добавлена.')

#             return redirect(f'{request.path}?level={level}')

#         elif action == 'delete':
#             obj_id = request.POST.get('id')
#             if level == 'region':
#                 obj = get_object_or_404(Region, id=obj_id)
#                 obj.delete()
#                 messages.success(request, f'Район "{obj.name}" удалён.')
#             elif level == 'substation':
#                 obj = get_object_or_404(Substation, id=obj_id)
#                 obj.delete()
#                 messages.success(request, f'Подстанция "{obj.name}" удалена.')
#             elif level == 'feeder':
#                 obj = get_object_or_404(Feeder, id=obj_id)
#                 obj.delete()
#                 messages.success(request, f'Фидер "{obj.name}" удалён.')
#             elif level == 'tp':
#                 obj = get_object_or_404(TransformerSubstation, id=obj_id)
#                 obj.delete()
#                 messages.success(request, f'ТП "{obj.name}" удалена.')
#             return redirect(f'{request.path}?level={level}')

#     # GET – получаем списки объектов для текущего уровня
#     if level == 'region':
#         objects = Region.objects.all().order_by('name')
#         parent_field = None
#     elif level == 'substation':
#         objects = Substation.objects.select_related('region').all().order_by('name')
#         parent_field = 'region'
#     elif level == 'feeder':
#         objects = Feeder.objects.select_related('substation').all().order_by('name')
#         parent_field = 'substation'
#     elif level == 'tp':
#         objects = TransformerSubstation.objects.select_related('feeder').all().order_by('name')
#         parent_field = 'feeder'
#     else:
#         objects = []
#         parent_field = None

#     # Формируем заголовки для таблицы
#     if level == 'region':
#         columns = ['Название', 'Головной счётчик']
#         fields = ['name', 'head_meter']
#     elif level == 'substation':
#         columns = ['Название', 'Район', 'Головной счётчик']
#         fields = ['name', 'region', 'head_meter']
#     elif level == 'feeder':
#         columns = ['Название', 'Подстанция', 'Головной счётчик']
#         fields = ['name', 'substation', 'head_meter']
#     elif level == 'tp':
#         columns = ['Название', 'Фидер', 'Головной счётчик']
#         fields = ['name', 'feeder', 'head_meter']
#     else:
#         columns = []
#         fields = []

#     context = {
#         'level': level,
#         'objects': objects,
#         'columns': columns,
#         'fields': fields,
#         'regions': regions,
#         'substations_all': substations_all,
#         'feeders_all': feeders_all,
#         'tps_all': tps_all,
#         'devices': devices,
#     }
#     return render(request, 'meters/directories.html', context)

# @login_required
# @user_passes_test(is_allowed_user)
# def directories_table(request):
#     """Возвращает только таблицу для указанного уровня (AJAX)."""
#     level = request.GET.get('level', 'region')

#     # Получаем списки объектов для текущего уровня (аналогично directories)
#     if level == 'region':
#         objects = Region.objects.all().order_by('name')
#         columns = ['Название', 'Головной счётчик']
#         fields = ['name', 'head_meter']
#         parent_field = None
#     elif level == 'substation':
#         objects = Substation.objects.select_related('region').all().order_by('name')
#         columns = ['Название', 'Район', 'Головной счётчик']
#         fields = ['name', 'region', 'head_meter']
#         parent_field = 'region'
#     elif level == 'feeder':
#         objects = Feeder.objects.select_related('substation').all().order_by('name')
#         columns = ['Название', 'Подстанция', 'Головной счётчик']
#         fields = ['name', 'substation', 'head_meter']
#         parent_field = 'substation'
#     elif level == 'tp':
#         objects = TransformerSubstation.objects.select_related('feeder').all().order_by('name')
#         columns = ['Название', 'Фидер', 'Головной счётчик']
#         fields = ['name', 'feeder', 'head_meter']
#         parent_field = 'feeder'
#     else:
#         objects = []
#         columns = []
#         fields = []

#     context = {
#         'level': level,
#         'objects': objects,
#         'columns': columns,
#         'fields': fields,
#     }
#     return render(request, 'meters/_directories_table.html', context)