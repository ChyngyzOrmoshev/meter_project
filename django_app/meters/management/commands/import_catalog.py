import os
import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction
from meters.models import MeterModel

class Command(BaseCommand):
    help = 'Import meter models from Catalog.xlsx'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            help='Path to Excel file (default: /app/Catalog.xlsx)',
            default='/app/Catalog.xlsx'
        )

    def handle(self, *args, **options):
        file_path = options['file']
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'File not found: {file_path}'))
            return

        self.stdout.write(f'📂 Reading {file_path}...')
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Определяем заголовки (первая строка)
        headers = [cell.value for cell in ws[1]]
        # Ожидаемые колонки:
        # Код, Наименование, Значность, Фазность, Номинальный ток, Номинальное напряжение,
        # Тип прибора учета, Период, Тип АСКУЭ, Для API
        # Если порядок другой, можно адаптировать, но обычно он такой же как в Catalog.xlsx.

        # Создаём маппинг позиций колонок
        col_map = {}
        for idx, name in enumerate(headers):
            if name:
                col_map[name.strip()] = idx

        required = ['Код', 'Наименование', 'Значность', 'Фазность', 'Номинальный ток',
                    'Номинальное напряжение', 'Тип прибора учета', 'Период', 'Тип АСКУЭ', 'Для API']
        for r in required:
            if r not in col_map:
                self.stdout.write(self.style.ERROR(f'Column "{r}" not found in Excel file'))
                return

        created = 0
        updated = 0
        errors = 0

        with transaction.atomic():
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not any(row):
                    continue
                try:
                    # Извлекаем значения
                    catalog_code = str(row[col_map['Код']]).strip() if row[col_map['Код']] else ''
                    model_name = str(row[col_map['Наименование']]).strip() if row[col_map['Наименование']] else ''
                    digit_capacity = str(row[col_map['Значность']]).strip() if row[col_map['Значность']] else ''
                    phases_text = str(row[col_map['Фазность']]).strip() if row[col_map['Фазность']] else ''
                    phases = 1 if 'одно' in phases_text.lower() else 3
                    nominal_current = str(row[col_map['Номинальный ток']]).strip() if row[col_map['Номинальный ток']] else ''
                    nominal_voltage = str(row[col_map['Номинальное напряжение']]).strip() if row[col_map['Номинальное напряжение']] else ''
                    system_type = str(row[col_map['Тип прибора учета']]).strip() if row[col_map['Тип прибора учета']] else ''
                    period = str(row[col_map['Период']]).strip() if row[col_map['Период']] else ''
                    device_type_id = str(row[col_map['Тип АСКУЭ']]).strip() if row[col_map['Тип АСКУЭ']] else ''
                    device_type_str = str(row[col_map['Для API']]).strip() if row[col_map['Для API']] else ''

                    if not model_name:
                        continue

                    obj, is_created = MeterModel.objects.update_or_create(
                        catalog_code=catalog_code,   # теперь ключ — код
                        defaults={
                            'model_name': model_name,
                            'digit_capacity': digit_capacity,
                            'phases': phases,
                            'nominal_current': nominal_current,
                            'nominal_voltage': nominal_voltage,
                            'system_type': system_type,
                            'period': period,
                            'device_type_id': device_type_id,
                            'device_type_str': device_type_str,
                        }
                    )
                    if is_created:
                        created += 1
                    else:
                        updated += 1
                except Exception as e:
                    errors += 1
                    self.stdout.write(self.style.ERROR(f'Error on row {row}: {e}'))

        self.stdout.write(self.style.SUCCESS(
            f'✅ Импорт завершён.\n'
            f'   Создано: {created}\n'
            f'   Обновлено: {updated}\n'
            f'   Ошибок: {errors}'
        ))